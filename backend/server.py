from fastapi import FastAPI, APIRouter, HTTPException, Query, Response, Request, Depends
from fastapi.responses import FileResponse, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any, Tuple
import uuid
from datetime import datetime, timezone, timedelta
import asyncio
import resend
import bcrypt
import jwt as pyjwt
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from unidecode import unidecode
import io
import base64
from openpyxl import Workbook
import openpyxl.utils
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from Levenshtein import ratio as levenshtein_ratio
import qrcode

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend configuration
resend.api_key = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

# PDF storage directory
PDF_DIR = ROOT_DIR / 'pdfs'
PDF_DIR.mkdir(exist_ok=True)

# Company info
COMPANY_INFO = {
    "name": "DCLIC INFORMATIQUE",
    "address": "30 AVENUE DU GENERAL DE GAULLE, 19140 UZERCHE",
    "phone": "05.55.73.57.20",
    "email": "contact@d-clic-informatique.fr"
}

# Conditions de réparation (mises à jour avec RGPD)
CONDITIONS_REPARATION = {
    "prise_en_charge": (
        "Dclic Informatique est responsable du matériel confié en cas de détérioration "
        "(partielle ou totale) ou de vol survenu pendant sa prise en charge. "
        "Le matériel est pris en charge en l'état. Dclic Informatique ne pourra être tenu "
        "responsable d'une aggravation d'une panne préexistante. "
        "La société ne pourra être tenue responsable en cas de perte de données. Le client "
        "reconnaît être seul responsable de la sauvegarde de ses données et accepte le risque "
        "de perte totale ou partielle lors de l'intervention. "
        "En cas de réinstallation du système, le client devra fournir une licence Windows "
        "valide et officielle. À défaut, une version d'évaluation limitée à 30 jours pourra "
        "être installée, dans la mesure du possible."
    ),
    "delais": (
        "Les délais de réparation dépendent de la charge de travail et de la disponibilité "
        "des pièces détachées. Aucune indemnité ne pourra être demandée en cas de dépassement "
        "des délais annoncés. Le client peut, s'il le souhaite, souscrire à l'option "
        "Réparation urgente (25 €), permettant une prise en charge prioritaire."
    ),
    "devis": (
        "Tout devis est gratuit s'il est accepté. En cas de refus, un montant forfaitaire "
        "de 15 € TTC sera facturé."
    ),
    "tarifs": (
        "Le forfait de réparation en atelier est fixé à 63 € TTC. "
        "En cas de panne matérielle nécessitant le remplacement de pièces, le client sera "
        "contacté et informé du coût des réparations. Aucune intervention ne sera réalisée "
        "sans son accord préalable."
    ),
    "reglement": (
        "Le règlement s'effectue au moment de la restitution du matériel, directement en boutique."
    ),
    "garantie": (
        "Les réparations sont garanties :\n"
        "• 3 mois pour la main-d'œuvre\n"
        "• 1 an pour les pièces remplacées (sauf indication contraire sur la facture)\n"
        "La garantie s'applique uniquement à la même panne et ne couvre pas automatiquement "
        "des symptômes similaires. La garantie ne couvre pas les dommages résultant d'une "
        "mauvaise utilisation, d'un choc, d'une surtension ou d'une intervention extérieure."
    ),
    "abandon": (
        "Tout appareil non récupéré dans un délai de 6 mois et 1 jour sera considéré comme "
        "abandonné. Le client sera contacté au préalable. En l'absence de réponse, Dclic "
        "Informatique se réserve le droit de disposer librement du matériel. "
        "Une proposition de reprise du matériel peut être faite au client avant la sortie "
        "de l'appareil des locaux. Le règlement s'effectuera alors exclusivement par chèque, "
        "sur présentation d'un justificatif d'identité et de domicile."
    ),
    "contestations": (
        "En confiant son matériel à Dclic Informatique, le client reconnaît avoir pris "
        "connaissance et accepté les présentes conditions. En cas de litige, seul le "
        "Tribunal de Commerce de Brive-La-Gaillarde sera compétent."
    ),
    "donnees_personnelles": (
        "Les informations collectées sont nécessaires à la gestion des réparations et à la "
        "relation client. Elles sont utilisées uniquement dans ce cadre et ne sont en aucun "
        "cas transmises à des tiers sans consentement. Conformément à la réglementation en "
        "vigueur, le client dispose d'un droit d'accès, de rectification et de suppression "
        "de ses données, qu'il peut exercer sur simple demande auprès de Dclic Informatique. "
        "Consultez notre politique de confidentialité pour plus d'informations."
    ),
}

# Politique de confidentialité (RGPD)
PRIVACY_POLICY = {
    "title": "Politique de confidentialité",
    "sections": [
        {"title": "1. Responsable du traitement", "content": "Les données personnelles collectées sont traitées par :\n\nDclic Informatique\n30 avenue du Général De Gaulle\n19140 UZERCHE"},
        {"title": "2. Données collectées", "content": "Dans le cadre de l'activité de réparation informatique, les données suivantes peuvent être collectées :\n\n• Nom et prénom\n• Numéro de téléphone\n• Adresse email\n• Informations relatives au matériel confié\n• Signature du client"},
        {"title": "3. Finalité du traitement", "content": "Ces données sont collectées uniquement pour :\n\n• la gestion des réparations\n• la communication avec le client\n• l'édition de documents (fiches, factures, etc.)\n\nElles ne sont en aucun cas utilisées à des fins commerciales sans consentement."},
        {"title": "4. Conservation des données", "content": "Les données sont conservées pendant la durée nécessaire à la gestion des réparations et au suivi client, puis archivées."},
        {"title": "5. Partage des données", "content": "Les données ne sont pas transmises à des tiers, sauf obligation légale ou nécessité technique liée à la réparation."},
        {"title": "6. Sécurité", "content": "Dclic Informatique met en œuvre des mesures techniques et organisationnelles pour protéger les données contre tout accès non autorisé, perte ou divulgation."},
        {"title": "7. Droits du client", "content": "Conformément à la réglementation en vigueur, le client dispose des droits suivants :\n\n• droit d'accès\n• droit de rectification\n• droit de suppression\n\nToute demande peut être adressée directement à Dclic Informatique."},
        {"title": "8. Acceptation", "content": "En confiant son matériel, le client reconnaît avoir pris connaissance de la présente politique de confidentialité."},
    ],
}

# Matériel fourni options
MATERIEL_OPTIONS = [
    "pc_portable", "pc_fixe", "sacoche", "imprimante", "chargeur_pc", "disque_dur_externe",
    "souris", "webcam", "cd_dvd", "clavier", "cle_usb", "cables_divers",
    "cle_wifi", "ecran", "onduleur", "enceintes", "documents_divers", "ipad"
]

# Statuts client (simplifié pour suivi public)
STATUTS_CLIENT = [
    "Réparation enregistrée",
    "En cours de diagnostic", 
    "En attente pièce/intervention",
    "En cours de réparation",
    "Appareil prêt"
]

# Statuts commande
STATUTS_COMMANDE = [
    "En attente de commande",
    "Commandé",
    "En attente réception",
    "Reçu",
    "Livré/Récupéré",
    "Réglé",
    "Annulé"
]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ===================== MODELS =====================

# Client Models
class ClientBase(BaseModel):
    nom: str
    prenom: str
    telephone: str
    telephone2: Optional[str] = None
    email: Optional[EmailStr] = None
    adresse: Optional[str] = None

class ClientCreate(ClientBase):
    pass

class ClientUpdate(BaseModel):
    nom: Optional[str] = None
    prenom: Optional[str] = None
    telephone: Optional[str] = None
    telephone2: Optional[str] = None
    email: Optional[EmailStr] = None
    adresse: Optional[str] = None

class Client(ClientBase):
    id: str
    created_at: str
    updated_at: str

# Repair Models - Extended
class ReparationBase(BaseModel):
    client_id: str
    # Matériel fourni (cases à cocher)
    materiel_fourni: Optional[Dict[str, bool]] = None
    autre_materiel: Optional[str] = None
    # Forfaits & options (cases à cocher — tarif 2026)
    # Clés possibles : express_10, rapide_30, standard_63, urgence_89, apple_89,
    # imprimante_45, recup_sain_63, recup_defectueux_79, sauvegarde_10, devis_15
    forfaits: Optional[List[str]] = None
    # Technique
    mot_de_passe: Optional[str] = None
    description_panne: str
    observations_client: Optional[str] = None
    # Protection juridique : N° série + état du matériel à la prise en charge
    numero_serie: Optional[str] = None
    etat_depot: Optional[str] = None
    # Diagnostic et action
    diagnostic: Optional[str] = None
    action_realisee: Optional[str] = None
    # Conseils au client (affichés sur la fiche compte-rendu)
    conseils: Optional[str] = None
    prix: Optional[float] = None
    # Statuts
    statut: str = "Réparation enregistrée"
    statut_interne: str = "En cours"

class ReparationCreate(ReparationBase):
    pass

class ReparationUpdate(BaseModel):
    client_id: Optional[str] = None
    materiel_fourni: Optional[Dict[str, bool]] = None
    autre_materiel: Optional[str] = None
    forfaits: Optional[List[str]] = None
    mot_de_passe: Optional[str] = None
    description_panne: Optional[str] = None
    observations_client: Optional[str] = None
    numero_serie: Optional[str] = None
    etat_depot: Optional[str] = None
    diagnostic: Optional[str] = None
    action_realisee: Optional[str] = None
    conseils: Optional[str] = None
    prix: Optional[float] = None
    statut: Optional[str] = None
    statut_interne: Optional[str] = None

class Reparation(BaseModel):
    id: str
    numero: str
    client_id: str
    tracking_id: Optional[str] = None
    date_creation: str
    heure_creation: Optional[str] = None
    date_modification: str
    # Legacy field support
    marque: Optional[str] = None
    modele: Optional[str] = None
    probleme_declare: Optional[str] = None
    # New fields
    materiel_fourni: Optional[Dict[str, bool]] = None
    autre_materiel: Optional[str] = None
    forfaits: Optional[List[str]] = None
    urgence: Optional[bool] = False  # legacy, déduit de forfaits contenant "urgence_89"
    mot_de_passe: Optional[str] = None
    description_panne: Optional[str] = None
    observations_client: Optional[str] = None
    numero_serie: Optional[str] = None
    etat_depot: Optional[str] = None
    diagnostic: Optional[str] = None
    action_realisee: Optional[str] = None
    conseils: Optional[str] = None
    prix: Optional[float] = None
    statut: str = "Réparation enregistrée"
    statut_interne: Optional[str] = "En cours"
    # Client info
    client_nom: Optional[str] = None
    client_prenom: Optional[str] = None
    client_email: Optional[str] = None
    client_telephone: Optional[str] = None
    # Signature client
    signature_b64: Optional[str] = None
    date_signature: Optional[str] = None
    nom_signataire: Optional[str] = None
    envoye_sans_signature: Optional[bool] = False
    # Encaissement
    encaissement_id: Optional[str] = None
    date_paiement: Optional[str] = None

# Commande Models
class CommandeBase(BaseModel):
    client_id: str
    reference_produit: Optional[str] = None
    designation: str
    fournisseur: Optional[str] = None
    quantite: int = 1
    prix_achat: Optional[float] = None
    prix_vente: Optional[float] = None
    statut: str = "En attente de commande"
    remarques: Optional[str] = None

class CommandeCreate(CommandeBase):
    pass

class CommandeUpdate(BaseModel):
    client_id: Optional[str] = None
    reference_produit: Optional[str] = None
    designation: Optional[str] = None
    fournisseur: Optional[str] = None
    quantite: Optional[int] = None
    prix_achat: Optional[float] = None
    prix_vente: Optional[float] = None
    statut: Optional[str] = None
    remarques: Optional[str] = None

class Commande(CommandeBase):
    id: str
    numero: str
    date_creation: str
    date_modification: str
    montant_total: Optional[float] = None
    client_nom: Optional[str] = None
    client_prenom: Optional[str] = None
    client_telephone: Optional[str] = None

# Cash Register Models (Caisse)
class CaisseEntryBase(BaseModel):
    type: str  # "entree" or "sortie"
    montant: float
    description: str
    mode_paiement: Optional[str] = None
    reparation_id: Optional[str] = None
    client_id: Optional[str] = None

class CaisseEntryCreate(CaisseEntryBase):
    pass

class CaisseEntry(CaisseEntryBase):
    id: str
    date: str

# Encaissement Models (entrées uniquement)
# Types correspondant aux forfaits du tarif 2026
TYPES_RECETTE = {
    "standard_63": {"label": "Réparation standard", "ttc": 63.0, "ht": 52.50},
    "rapide_30": {"label": "Réparation rapide", "ttc": 30.0, "ht": 25.0},
    "express_10": {"label": "Réparation express", "ttc": 10.0, "ht": 8.33},
    "urgence_89": {"label": "Forfait urgence (24h)", "ttc": 89.0, "ht": 74.17},
    "apple_89": {"label": "Forfait Apple", "ttc": 89.0, "ht": 74.17},
    "imprimante_45": {"label": "Forfait imprimante", "ttc": 45.0, "ht": 37.50},
    "recup_sain_63": {"label": "Récup. données (support sain)", "ttc": 63.0, "ht": 52.50},
    "recup_defectueux_79": {"label": "Récup. données (support défectueux)", "ttc": 79.0, "ht": 65.83},
    "sauvegarde_10": {"label": "Option sauvegarde", "ttc": 10.0, "ht": 8.33},
    "devis_15": {"label": "Devis", "ttc": 15.0, "ht": 12.50},
    "ventes": {"label": "Ventes", "ttc": None, "ht": None},
    "autre": {"label": "Autre", "ttc": None, "ht": None},
    "mixte": {"label": "Mixte (plusieurs lignes)", "ttc": None, "ht": None},
    # Legacy
    "forfait_63": {"label": "Réparation standard", "ttc": 63.0, "ht": 52.50},
}

# Catalogue des forfaits applicables sur une fiche réparation
FORFAITS_CATALOG = [
    {"key": "express_10", "label": "Réparation express (<10 min)", "prix": 10.0},
    {"key": "rapide_30", "label": "Réparation rapide (<30 min)", "prix": 30.0},
    {"key": "standard_63", "label": "Réparation standard (>30 min)", "prix": 63.0},
    {"key": "urgence_89", "label": "Forfait urgence (dépannage 24h)", "prix": 89.0},
    {"key": "apple_89", "label": "Forfait Apple", "prix": 89.0},
    {"key": "imprimante_45", "label": "Forfait nettoyage imprimante", "prix": 45.0},
    {"key": "recup_sain_63", "label": "Récupération données — support sain", "prix": 63.0},
    {"key": "recup_defectueux_79", "label": "Récupération données — support défectueux", "prix": 79.0},
    {"key": "sauvegarde_10", "label": "Option sauvegarde", "prix": 10.0},
    {"key": "devis_15", "label": "Devis (offert si réparation acceptée)", "prix": 15.0},
]
FORFAIT_LABELS = {f["key"]: f["label"] for f in FORFAITS_CATALOG}

def format_forfaits_list(forfaits_keys: Optional[List[str]]) -> List[str]:
    """Transforme une liste de clés en labels lisibles."""
    if not forfaits_keys:
        return []
    return [FORFAIT_LABELS.get(k, k) for k in forfaits_keys]

class PaiementDetail(BaseModel):
    mode: str  # especes, cb, cheque, virement
    montant: float

class LigneRecette(BaseModel):
    """Ligne unitaire d'un encaissement multi-produits (ex : forfait + vente)."""
    type_recette: str
    montant_ttc: float
    montant_ht: Optional[float] = None
    description: Optional[str] = None

class EncaissementBase(BaseModel):
    type_recette: str
    montant_ttc: float
    montant_ht: Optional[float] = None
    paiements: List[PaiementDetail]  # Permet plusieurs modes de paiement
    lignes: Optional[List[LigneRecette]] = None  # Détail multi-produits (facultatif)
    client_id: Optional[str] = None
    reference: Optional[str] = None
    remarque: Optional[str] = None
    reparation_id: Optional[str] = None  # si créé depuis une fiche réparation

class EncaissementCreate(EncaissementBase):
    pass

class Encaissement(EncaissementBase):
    id: str
    date: str
    client_nom: Optional[str] = None
    client_prenom: Optional[str] = None

class DashboardStats(BaseModel):
    total_clients: int
    total_reparations: int
    reparations_en_cours: int
    reparations_terminees: int
    total_commandes: int
    commandes_en_attente: int
    total_caisse_jour: float
    total_entrees_jour: float
    total_sorties_jour: float

# ===================== HELPER FUNCTIONS =====================

async def get_next_repair_number():
    """Generate next repair number in format REP-YYYY-XXXX"""
    year = datetime.now(timezone.utc).year
    prefix = f"REP-{year}-"
    
    last_repair = await db.reparations.find_one(
        {"numero": {"$regex": f"^{prefix}"}},
        sort=[("numero", -1)]
    )
    
    if last_repair:
        try:
            last_num = int(last_repair["numero"].split("-")[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    return f"{prefix}{next_num:04d}"

async def get_next_commande_number():
    """Generate next commande number in format cmd-DD-MM-YYYY-XXXX"""
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%d-%m-%Y")
    prefix = f"cmd-{date_str}-"
    
    last_commande = await db.commandes.find_one(
        {"numero": {"$regex": f"^{prefix}"}},
        sort=[("numero", -1)]
    )
    
    if last_commande:
        try:
            last_num = int(last_commande["numero"].split("-")[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    return f"{prefix}{next_num:04d}"

def sanitize_filename(filename):
    """Remove problematic characters from filename"""
    filename = unidecode(filename)
    filename = filename.replace(" ", "_")
    filename = "".join(c for c in filename if c.isalnum() or c in "._-")
    return filename

def generate_qr_code(data: str) -> bytes:
    """Generate QR code as PNG bytes"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer.getvalue()

def get_materiel_fourni_list(materiel: Dict[str, bool], autre: str = None) -> List[str]:
    """Convert materiel dict to readable list"""
    labels = {
        "pc_portable": "PC portable",
        "pc_fixe": "PC fixe",
        "sacoche": "Sacoche",
        "imprimante": "Imprimante",
        "chargeur_pc": "Chargeur PC portable",
        "disque_dur_externe": "Disque dur externe",
        "souris": "Souris",
        "webcam": "Webcam",
        "cd_dvd": "CD/DVD divers",
        "clavier": "Clavier",
        "cle_usb": "Clé USB",
        "cables_divers": "Câbles divers",
        "cle_wifi": "Clé Wifi",
        "ecran": "Écran",
        "onduleur": "Onduleur",
        "enceintes": "Enceintes",
        "documents_divers": "Documents divers",
        "ipad": "iPad"
    }
    result = []
    if materiel:
        for key, checked in materiel.items():
            if checked and key in labels:
                result.append(labels[key])
    if autre:
        result.append(f"Autre: {autre}")
    return result

def _fr_date(iso_str: Optional[str]) -> str:
    """Convertit une date ISO (YYYY-MM-DD[T...]) en jj/mm/aaaa."""
    if not iso_str:
        return ""
    raw = iso_str[:10]
    try:
        return datetime.strptime(raw, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return raw

def generate_client_pdf(reparation: dict, client: dict, tracking_url: str = None, force_no_signature: bool = False) -> bytes:
    """Generate PDF for client (without password, with conditions)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           leftMargin=1.5*cm, rightMargin=1.5*cm,
                           topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
        fontSize=16, spaceAfter=10, textColor=colors.HexColor('#84CC16'), alignment=1)
    header_style = ParagraphStyle('Header', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748B'), alignment=1)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
        fontSize=11, spaceBefore=12, spaceAfter=6, textColor=colors.HexColor('#0F172A'))
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'],
        fontSize=9, spaceAfter=3)
    small_style = ParagraphStyle('Small', parent=styles['Normal'],
        fontSize=7, textColor=colors.HexColor('#64748B'), spaceAfter=2)
    condition_title = ParagraphStyle('ConditionTitle', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Bold', textColor=colors.HexColor('#0F172A'), spaceAfter=2)
    condition_text = ParagraphStyle('ConditionText', parent=styles['Normal'],
        fontSize=7, textColor=colors.HexColor('#374151'), spaceAfter=4)
    
    elements = []
    
    # Header
    elements.append(Paragraph(COMPANY_INFO["name"], title_style))
    elements.append(Paragraph(COMPANY_INFO["address"], header_style))
    elements.append(Paragraph(f"Tél: {COMPANY_INFO['phone']} | Email: {COMPANY_INFO['email']}", header_style))
    elements.append(Spacer(1, 15))
    
    # Title with number and date
    elements.append(Paragraph(f"<b>FICHE DE RÉPARATION N° {reparation['numero']}</b>", section_style))
    elements.append(Paragraph(f"Date: {_fr_date(reparation.get('date_creation'))} - Heure: {reparation.get('heure_creation', '')}", normal_style))
    if reparation.get('urgence'):
        elements.append(Paragraph("<font color='red'><b>⚠ RÉPARATION URGENTE (+25€)</b></font>", normal_style))
    elements.append(Spacer(1, 10))
    
    # Client info
    elements.append(Paragraph("<b>INFORMATIONS CLIENT</b>", section_style))
    client_data = [
        ["Nom:", f"{client.get('prenom', '')} {client.get('nom', '')}"],
        ["Téléphone:", client.get('telephone', '-')],
        ["Email:", client.get('email', '-') or '-'],
    ]
    client_table = Table(client_data, colWidths=[3*cm, 13*cm])
    client_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 10))
    
    # Matériel fourni
    materiel_list = get_materiel_fourni_list(
        reparation.get('materiel_fourni', {}),
        reparation.get('autre_materiel')
    )
    if materiel_list:
        elements.append(Paragraph("<b>MATÉRIEL FOURNI</b>", section_style))
        elements.append(Paragraph(", ".join(materiel_list), normal_style))
        elements.append(Spacer(1, 10))

    # Forfaits appliqués
    forfaits_labels = format_forfaits_list(reparation.get('forfaits'))
    if forfaits_labels:
        elements.append(Paragraph("<b>FORFAITS / OPTIONS</b>", section_style))
        elements.append(Paragraph(", ".join(forfaits_labels), normal_style))
        elements.append(Spacer(1, 10))

    # État du matériel à la prise en charge (protection juridique)
    numero_serie = reparation.get('numero_serie') or ''
    etat_depot = reparation.get('etat_depot') or ''
    if numero_serie or etat_depot:
        elements.append(Paragraph("<b>ÉTAT DU MATÉRIEL À LA PRISE EN CHARGE</b>", section_style))
        if numero_serie:
            elements.append(Paragraph(f"<b>N° de série :</b> {numero_serie}", normal_style))
        if etat_depot:
            elements.append(Paragraph(f"<b>Observations :</b> {etat_depot}", normal_style))
        elements.append(Spacer(1, 10))

    # Description
    elements.append(Paragraph("<b>DESCRIPTION DE LA PANNE</b>", section_style))
    elements.append(Paragraph(reparation.get('description_panne', '-'), normal_style))
    
    if reparation.get('observations_client'):
        elements.append(Spacer(1, 5))
        elements.append(Paragraph("<b>Observations du client:</b>", normal_style))
        elements.append(Paragraph(reparation.get('observations_client', ''), normal_style))
    elements.append(Spacer(1, 10))
    
    # Diagnostic et action
    if reparation.get('diagnostic') or reparation.get('action_realisee'):
        elements.append(Paragraph("<b>DIAGNOSTIC ET INTERVENTION</b>", section_style))
        if reparation.get('diagnostic'):
            elements.append(Paragraph(f"<b>Diagnostic:</b> {reparation.get('diagnostic')}", normal_style))
        if reparation.get('action_realisee'):
            elements.append(Paragraph(f"<b>Action réalisée:</b> {reparation.get('action_realisee')}", normal_style))
        elements.append(Spacer(1, 10))
    
    # Prix et statut
    elements.append(Paragraph("<b>STATUT ET TARIF</b>", section_style))
    status_data = [
        ["Statut:", reparation.get('statut', '-')],
    ]
    if reparation.get('prix'):
        status_data.append(["Prix:", f"{reparation['prix']:.2f} €"])
    status_table = Table(status_data, colWidths=[3*cm, 13*cm])
    status_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 15))
    
    # QR Code for tracking
    if tracking_url:
        elements.append(Paragraph("<b>SUIVI DE VOTRE RÉPARATION</b>", section_style))
        elements.append(Paragraph(f"Scannez le QR code ou visitez:", small_style))
        elements.append(Paragraph(f"<font color='blue'>{tracking_url}</font>", small_style))
        
        # Generate QR code
        qr_bytes = generate_qr_code(tracking_url)
        qr_img = Image(io.BytesIO(qr_bytes), width=2.5*cm, height=2.5*cm)
        elements.append(qr_img)
        elements.append(Spacer(1, 10))
    
    # Conditions de réparation
    elements.append(Paragraph("<b>CONDITIONS DE RÉPARATION</b>", section_style))
    elements.append(Paragraph("<b>Prise en charge du matériel:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["prise_en_charge"], condition_text))
    elements.append(Paragraph("<b>Délais:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["delais"], condition_text))
    elements.append(Paragraph("<b>Devis:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["devis"], condition_text))
    elements.append(Paragraph("<b>Tarifs:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["tarifs"], condition_text))
    elements.append(Paragraph("<b>Règlement:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["reglement"], condition_text))
    elements.append(Paragraph("<b>Garantie:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["garantie"], condition_text))
    elements.append(Paragraph("<b>Abandon:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["abandon"], condition_text))
    elements.append(Paragraph("<b>Contestations:</b>", condition_title))
    elements.append(Paragraph(CONDITIONS_REPARATION["contestations"], condition_text))
    
    elements.append(Spacer(1, 15))
    # Bloc signature
    if reparation.get("signature_b64"):
        signataire = reparation.get("nom_signataire") or f"{client.get('prenom','')} {client.get('nom','')}".strip()
        date_sig_iso = reparation.get("date_signature", "")[:10]
        # Format jj/mm/aaaa
        if date_sig_iso and len(date_sig_iso) == 10:
            try:
                date_sig = datetime.strptime(date_sig_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                date_sig = date_sig_iso
        else:
            date_sig = date_sig_iso
        elements.append(Paragraph("<b>Lu et approuvé, bon pour accord</b>", normal_style))
        elements.append(Paragraph(f"Nom : {signataire}    —    Date : {date_sig}", normal_style))
        elements.append(Spacer(1, 6))
        try:
            raw = reparation["signature_b64"]
            if raw.startswith("data:"):
                raw = raw.split(",", 1)[1]
            sig_bytes = base64.b64decode(raw)
            # Signature triplée : 15cm × 6cm (était 5cm × 2cm)
            sig_img = Image(io.BytesIO(sig_bytes), width=15 * cm, height=6 * cm)
            elements.append(sig_img)
        except Exception as exc:
            logger.error(f"Failed to embed signature: {exc}")
            elements.append(Paragraph("<i>[Signature non affichable]</i>", small_style))
    elif reparation.get("envoye_sans_signature") or force_no_signature:
        elements.append(Paragraph(
            "<font color='#DC2626'><b>Document envoyé sans signature du client</b></font>",
            normal_style,
        ))
    else:
        elements.append(Paragraph("Signature du client (lu et approuvé) :", normal_style))
        elements.append(Spacer(1, 20))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generate_compte_rendu_pdf(reparation: dict, client: dict, ad_banner_bytes: Optional[bytes] = None) -> bytes:
    """Fiche compte rendu — tout sur UNE seule page A4.
    Le bandeau publicitaire est dessiné directement sur le canvas au bas de la page
    (via onFirstPage), avec réservation d'un bottomMargin adapté pour que le texte
    ne déborde jamais dessus.
    """
    buffer = io.BytesIO()

    # Préparation du bandeau (sanitization + dimensions finales) AVANT doc
    banner_draw = None
    banner_h = 0
    if ad_banner_bytes:
        try:
            from PIL import Image as PILImage
            pil = PILImage.open(io.BytesIO(ad_banner_bytes))
            pil.verify()
            pil = PILImage.open(io.BytesIO(ad_banner_bytes))
            if pil.mode in ("P", "RGBA"):
                pil = pil.convert("RGB")
            clean = io.BytesIO()
            pil.save(clean, format="JPEG", quality=85)
            clean.seek(0)
            iw, ih = pil.size
            max_w = 17 * cm
            max_h = 5 * cm  # bandeau plus compact pour garantir 1 page
            ratio = min(max_w / iw, max_h / ih)
            banner_w = iw * ratio
            banner_h = ih * ratio
            banner_draw = (clean, banner_w, banner_h)
        except Exception as exc:
            logger.error(f"Failed to prepare ad banner: {exc}")

    # Marge basse = hauteur bandeau + petit padding (si pas de bandeau, marge standard)
    bottom_margin = (banner_h + 0.8 * cm) if banner_h else 1.2 * cm

    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.2 * cm, bottomMargin=bottom_margin)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontSize=16, spaceAfter=4,
                                 textColor=colors.HexColor('#84CC16'), alignment=1)
    header_style = ParagraphStyle('Header', parent=styles['Normal'],
                                  fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                   fontSize=10.5, spaceBefore=8, spaceAfter=3,
                                   textColor=colors.HexColor('#0F172A'))
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'],
                                  fontSize=9.5, spaceAfter=2, leading=12)
    thank_style = ParagraphStyle('Thank', parent=styles['Normal'],
                                 fontSize=10, fontName='Helvetica-Oblique',
                                 textColor=colors.HexColor('#0F172A'), alignment=1, spaceAfter=4)
    tip_style = ParagraphStyle('Tip', parent=styles['Normal'],
                               fontSize=9.5, textColor=colors.HexColor('#374151'),
                               leading=13, spaceAfter=2)

    elements = []

    # En-tête
    elements.append(Paragraph(COMPANY_INFO["name"], title_style))
    elements.append(Paragraph(COMPANY_INFO["address"], header_style))
    elements.append(Paragraph(
        f"Tél : {COMPANY_INFO['phone']} | Email : {COMPANY_INFO['email']}",
        header_style))
    elements.append(Spacer(1, 6))

    elements.append(Paragraph(
        f"<b>COMPTE RENDU DE RÉPARATION — N° {reparation['numero']}</b>",
        section_style))
    elements.append(Paragraph(
        f"Date de dépôt : {_fr_date(reparation.get('date_creation'))} — "
        f"Date d'édition : {datetime.now().strftime('%d/%m/%Y')}",
        normal_style))
    elements.append(Spacer(1, 3))

    # Remerciement
    elements.append(Paragraph(
        f"Bonjour {client.get('prenom', '')} {client.get('nom', '')},<br/>"
        "Merci pour votre confiance. Voici le récapitulatif de l'intervention réalisée.",
        thank_style))

    # Client + matériel (compact)
    materiel_list = get_materiel_fourni_list(
        reparation.get('materiel_fourni', {}),
        reparation.get('autre_materiel'))
    info_data = [
        ["Client :", f"{client.get('prenom', '')} {client.get('nom', '')}"],
        ["Téléphone :", client.get('telephone', '-') or '-'],
        ["Matériel :", ", ".join(materiel_list) if materiel_list else "-"],
    ]
    if reparation.get('numero_serie'):
        info_data.append(["N° de série :", reparation.get('numero_serie', '-')])
    info_table = Table(info_data, colWidths=[3.5 * cm, 14.5 * cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
    ]))
    elements.append(info_table)

    # Ce qui a été fait
    elements.append(Paragraph("<b>PROBLÈME SIGNALÉ</b>", section_style))
    elements.append(Paragraph(reparation.get('description_panne') or '-', normal_style))

    if reparation.get('diagnostic'):
        elements.append(Paragraph("<b>DIAGNOSTIC</b>", section_style))
        elements.append(Paragraph(reparation['diagnostic'], normal_style))

    if reparation.get('action_realisee'):
        elements.append(Paragraph("<b>INTERVENTION RÉALISÉE</b>", section_style))
        elements.append(Paragraph(reparation['action_realisee'], normal_style))

    # Prix
    prix = reparation.get('prix')
    if prix is not None:
        prix_table = Table(
            [["Montant total de l'intervention :", f"{float(prix):.2f} €"]],
            colWidths=[12 * cm, 6 * cm],
        )
        prix_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 10.5),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 13),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#84CC16')),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
        ]))
        elements.append(Spacer(1, 6))
        elements.append(prix_table)

    # Conseils
    conseils = (reparation.get('conseils') or '').strip()
    if conseils:
        elements.append(Paragraph("<b>NOS CONSEILS</b>", section_style))
        tip_table = Table([[Paragraph(conseils.replace('\n', '<br/>'), tip_style)]],
                          colWidths=[18 * cm])
        tip_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F7FEE7')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#84CC16')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tip_table)

    # Message de fin
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "Nous restons à votre disposition pour toute question concernant cette intervention. "
        "N'hésitez pas à nous recommander auprès de vos proches !",
        normal_style))

    # Callback canvas : dessine le bandeau publicitaire en bas, fixe
    def _draw_banner(canvas, _doc):
        if not banner_draw:
            return
        clean_buf, bw, bh = banner_draw
        clean_buf.seek(0)
        page_w, _ = A4
        x = (page_w - bw) / 2
        y = 0.6 * cm  # 6 mm du bord inférieur
        try:
            canvas.drawImage(
                ImageReader(clean_buf), x, y,
                width=bw, height=bh,
                preserveAspectRatio=True, mask='auto')
        except Exception as exc:
            logger.error(f"Failed to draw banner on canvas: {exc}")

    doc.build(elements, onFirstPage=_draw_banner, onLaterPages=_draw_banner)
    buffer.seek(0)
    return buffer.getvalue()



def generate_internal_pdf(reparation: dict, client: dict) -> bytes:
    """Generate internal PDF (with password and all details)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=1.5*cm, rightMargin=1.5*cm,
                           topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
        fontSize=16, spaceAfter=10, textColor=colors.HexColor('#0F172A'), alignment=1)
    internal_badge = ParagraphStyle('InternalBadge', parent=styles['Normal'],
        fontSize=10, textColor=colors.white, backColor=colors.HexColor('#DC2626'),
        alignment=1, spaceBefore=5, spaceAfter=10)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
        fontSize=11, spaceBefore=12, spaceAfter=6, textColor=colors.HexColor('#0F172A'))
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'],
        fontSize=9, spaceAfter=3)
    
    elements = []
    
    # Internal badge
    elements.append(Paragraph("*** DOCUMENT INTERNE - NE PAS TRANSMETTRE AU CLIENT ***", internal_badge))
    
    # Header
    elements.append(Paragraph(f"FICHE INTERNE N° {reparation['numero']}", title_style))
    elements.append(Paragraph(f"Date: {_fr_date(reparation.get('date_creation'))} - Heure: {reparation.get('heure_creation', '')}", normal_style))
    if reparation.get('urgence'):
        elements.append(Paragraph("<font color='red'><b>⚠ RÉPARATION URGENTE</b></font>", normal_style))
    elements.append(Spacer(1, 10))
    
    # Client info
    elements.append(Paragraph("<b>CLIENT</b>", section_style))
    client_data = [
        ["Nom:", f"{client.get('prenom', '')} {client.get('nom', '')}"],
        ["Téléphone:", client.get('telephone', '-')],
        ["Email:", client.get('email', '-') or '-'],
        ["Adresse:", client.get('adresse', '-') or '-'],
    ]
    client_table = Table(client_data, colWidths=[3*cm, 13*cm])
    client_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 10))
    
    # Matériel fourni
    materiel_list = get_materiel_fourni_list(
        reparation.get('materiel_fourni', {}),
        reparation.get('autre_materiel')
    )
    if materiel_list:
        elements.append(Paragraph("<b>MATÉRIEL FOURNI</b>", section_style))
        elements.append(Paragraph(", ".join(materiel_list), normal_style))
        elements.append(Spacer(1, 10))

    # Forfaits appliqués
    forfaits_labels = format_forfaits_list(reparation.get('forfaits'))
    if forfaits_labels:
        elements.append(Paragraph("<b>FORFAITS / OPTIONS</b>", section_style))
        elements.append(Paragraph(", ".join(forfaits_labels), normal_style))
        elements.append(Spacer(1, 10))

    # MOT DE PASSE (visible only on internal) + N° série
    elements.append(Paragraph("<b>INFORMATIONS TECHNIQUES</b>", section_style))
    tech_data = [
        ["Mot de passe:", reparation.get('mot_de_passe', '-') or '-'],
        ["N° de série:", reparation.get('numero_serie', '-') or '-'],
    ]
    tech_table = Table(tech_data, colWidths=[3*cm, 13*cm])
    tech_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#E5E7EB')),
    ]))
    elements.append(tech_table)
    elements.append(Spacer(1, 10))

    # État à la prise en charge
    if reparation.get('etat_depot'):
        elements.append(Paragraph("<b>ÉTAT DU MATÉRIEL À LA PRISE EN CHARGE</b>", section_style))
        elements.append(Paragraph(reparation.get('etat_depot', ''), normal_style))
        elements.append(Spacer(1, 10))
    
    # Description
    elements.append(Paragraph("<b>DESCRIPTION DE LA PANNE</b>", section_style))
    elements.append(Paragraph(reparation.get('description_panne', '-'), normal_style))
    
    if reparation.get('observations_client'):
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(f"<b>Observations client:</b> {reparation.get('observations_client')}", normal_style))
    elements.append(Spacer(1, 10))
    
    # Diagnostic et action
    elements.append(Paragraph("<b>DIAGNOSTIC ET INTERVENTION</b>", section_style))
    elements.append(Paragraph(f"<b>Diagnostic:</b> {reparation.get('diagnostic', '-') or '-'}", normal_style))
    elements.append(Paragraph(f"<b>Action réalisée:</b> {reparation.get('action_realisee', '-') or '-'}", normal_style))
    elements.append(Spacer(1, 10))
    
    # Status
    elements.append(Paragraph("<b>STATUT</b>", section_style))
    status_data = [
        ["Statut client:", reparation.get('statut', '-')],
        ["Statut interne:", reparation.get('statut_interne', '-')],
        ["Prix:", f"{reparation.get('prix', 0):.2f} €" if reparation.get('prix') else '-'],
    ]
    status_table = Table(status_data, colWidths=[3*cm, 13*cm])
    status_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(status_table)
    
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Créé le: {_fr_date(reparation.get('date_creation'))}", normal_style))
    elements.append(Paragraph(f"Modifié le: {reparation['date_modification']}", normal_style))
    elements.append(Paragraph(f"Tracking ID: {reparation.get('tracking_id', '-')}", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# ===================== API ROUTES =====================

@api_router.get("/")
async def root():
    return {"message": "DCLIC Informatique API", "status": "running"}

# Dashboard
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    """Get dashboard statistics"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    total_clients = await db.clients.count_documents({})
    total_reparations = await db.reparations.count_documents({})
    reparations_en_cours = await db.reparations.count_documents({"statut_interne": "En cours"})
    reparations_terminees = await db.reparations.count_documents({"statut_interne": "Terminé"})
    total_commandes = await db.commandes.count_documents({})
    commandes_en_attente = await db.commandes.count_documents({"statut": {"$nin": ["Livré/Récupéré", "Réglé", "Annulé"]}})
    
    # Cash register for today
    today_entries = await db.caisse.find(
        {"date": {"$regex": f"^{today}"}},
        {"_id": 0}
    ).to_list(1000)
    
    total_entrees = sum(e["montant"] for e in today_entries if e["type"] == "entree")
    total_sorties = sum(e["montant"] for e in today_entries if e["type"] == "sortie")
    
    # Also add encaissements
    today_encaissements = await db.encaissements.find(
        {"date": {"$regex": f"^{today}"}},
        {"_id": 0}
    ).to_list(1000)
    total_entrees += sum(e.get("montant_ttc", e.get("montant", 0)) or 0 for e in today_encaissements)
    
    return DashboardStats(
        total_clients=total_clients,
        total_reparations=total_reparations,
        reparations_en_cours=reparations_en_cours,
        reparations_terminees=reparations_terminees,
        total_commandes=total_commandes,
        commandes_en_attente=commandes_en_attente,
        total_caisse_jour=total_entrees - total_sorties,
        total_entrees_jour=total_entrees,
        total_sorties_jour=total_sorties
    )

# ===================== CLIENTS CRUD =====================

@api_router.post("/clients", response_model=Client)
async def create_client(client: ClientCreate):
    """Create a new client"""
    now = datetime.now(timezone.utc).isoformat()
    client_doc = {
        "id": str(uuid.uuid4()),
        "nom": client.nom,
        "prenom": client.prenom,
        "telephone": client.telephone,
        "telephone2": client.telephone2,
        "email": client.email,
        "adresse": client.adresse,
        "created_at": now,
        "updated_at": now
    }
    await db.clients.insert_one(client_doc)
    client_doc.pop("_id", None)
    return client_doc

@api_router.get("/clients", response_model=List[Client])
async def get_clients(
    search: Optional[str] = Query(None, description="Search by name or phone"),
    limit: int = Query(100, le=500)
):
    """Get all clients with optional search"""
    if search:
        all_clients = await db.clients.find({}, {"_id": 0}).to_list(500)
        search_lower = search.lower()
        search_normalized = unidecode(search_lower)
        
        matched_clients = []
        for c in all_clients:
            nom_normalized = unidecode(c.get("nom", "").lower())
            prenom_normalized = unidecode(c.get("prenom", "").lower())
            full_name = f"{prenom_normalized} {nom_normalized}"
            telephone = c.get("telephone", "")
            
            if search in telephone:
                matched_clients.append(c)
                continue
            
            if (levenshtein_ratio(search_normalized, nom_normalized) > 0.7 or
                levenshtein_ratio(search_normalized, prenom_normalized) > 0.7 or
                levenshtein_ratio(search_normalized, full_name) > 0.6 or
                search_normalized in nom_normalized or
                search_normalized in prenom_normalized):
                matched_clients.append(c)
        
        return matched_clients[:limit]
    
    clients = await db.clients.find({}, {"_id": 0}).to_list(limit)
    return clients

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str):
    """Get a specific client"""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return client

@api_router.get("/clients/{client_id}/reparations")
async def get_client_reparations(client_id: str):
    """Get all repairs for a specific client"""
    reparations = await db.reparations.find({"client_id": client_id}, {"_id": 0}).sort("date_creation", -1).to_list(100)
    return reparations

@api_router.get("/clients/{client_id}/commandes")
async def get_client_commandes(client_id: str):
    """Get all orders for a specific client"""
    commandes = await db.commandes.find({"client_id": client_id}, {"_id": 0}).sort("date_creation", -1).to_list(100)
    return commandes

@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, update: ClientUpdate):
    """Update a client"""
    existing = await db.clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return updated

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str):
    """Delete a client"""
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return {"message": "Client supprimé"}

# ===================== REPARATIONS CRUD =====================

@api_router.get("/reparations/statuts-client")
async def get_statuts_client():
    """Get available client statuses"""
    return {"statuts": STATUTS_CLIENT}

@api_router.get("/reparations/materiel-options")
async def get_materiel_options():
    """Get available equipment options"""
    labels = {
        "pc_portable": "PC portable", "pc_fixe": "PC fixe", "sacoche": "Sacoche",
        "imprimante": "Imprimante", "chargeur_pc": "Chargeur PC portable",
        "disque_dur_externe": "Disque dur externe", "souris": "Souris", "webcam": "Webcam",
        "cd_dvd": "CD/DVD divers", "clavier": "Clavier", "cle_usb": "Clé USB",
        "cables_divers": "Câbles divers", "cle_wifi": "Clé Wifi", "ecran": "Écran",
        "onduleur": "Onduleur", "enceintes": "Enceintes", "documents_divers": "Documents divers",
        "ipad": "iPad"
    }
    return {"options": labels}

@api_router.post("/reparations", response_model=Reparation)
async def create_reparation(reparation: ReparationCreate):
    """Create a new repair"""
    client = await db.clients.find_one({"id": reparation.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    now = datetime.now(timezone.utc)
    numero = await get_next_repair_number()
    tracking_id = str(uuid.uuid4())[:8].upper()
    
    rep_doc = {
        "id": str(uuid.uuid4()),
        "numero": numero,
        "tracking_id": tracking_id,
        "client_id": reparation.client_id,
        "materiel_fourni": reparation.materiel_fourni or {},
        "autre_materiel": reparation.autre_materiel,
        "forfaits": reparation.forfaits or [],
        "urgence": "urgence_89" in (reparation.forfaits or []),
        "mot_de_passe": reparation.mot_de_passe,
        "description_panne": reparation.description_panne,
        "observations_client": reparation.observations_client,
        "numero_serie": reparation.numero_serie,
        "etat_depot": reparation.etat_depot,
        "diagnostic": reparation.diagnostic,
        "action_realisee": reparation.action_realisee,
        "conseils": reparation.conseils,
        "prix": reparation.prix,
        "statut": reparation.statut,
        "statut_interne": reparation.statut_interne,
        "date_creation": now.isoformat(),
        "heure_creation": now.strftime("%H:%M"),
        "date_modification": now.isoformat()
    }
    
    await db.reparations.insert_one(rep_doc)
    rep_doc.pop("_id", None)
    
    rep_doc["client_nom"] = client.get("nom")
    rep_doc["client_prenom"] = client.get("prenom")
    rep_doc["client_email"] = client.get("email")
    rep_doc["client_telephone"] = client.get("telephone")
    
    return rep_doc

@api_router.get("/reparations", response_model=List[Reparation])
async def get_reparations(
    search: Optional[str] = Query(None),
    statut: Optional[str] = Query(None),
    statut_interne: Optional[str] = Query(None),
    limit: int = Query(100, le=500)
):
    """Get all repairs with optional filters"""
    query = {}
    
    if statut:
        query["statut"] = statut
    if statut_interne:
        query["statut_interne"] = statut_interne
    
    reparations = await db.reparations.find(query, {"_id": 0}).sort("date_creation", -1).to_list(limit)
    
    for rep in reparations:
        # Migrate old data: if description_panne doesn't exist, use probleme_declare
        if "description_panne" not in rep and "probleme_declare" in rep:
            rep["description_panne"] = rep.get("probleme_declare", "")
        
        # Add missing fields for old records
        if "tracking_id" not in rep:
            rep["tracking_id"] = str(uuid.uuid4())[:8].upper()
        if "heure_creation" not in rep:
            rep["heure_creation"] = ""
        if "statut_interne" not in rep:
            rep["statut_interne"] = "Terminé" if rep.get("statut") == "Terminé" else "En cours"
        
        client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
        if client:
            rep["client_nom"] = client.get("nom")
            rep["client_prenom"] = client.get("prenom")
            rep["client_email"] = client.get("email")
            rep["client_telephone"] = client.get("telephone")
    
    if search:
        search_normalized = unidecode(search.lower())
        filtered = []
        for rep in reparations:
            client_name = f"{rep.get('client_prenom', '')} {rep.get('client_nom', '')}".lower()
            client_name_normalized = unidecode(client_name)
            numero = rep.get("numero", "").lower()
            
            if (search_normalized in client_name_normalized or
                search_normalized in numero or
                levenshtein_ratio(search_normalized, client_name_normalized) > 0.6):
                filtered.append(rep)
        
        return filtered
    
    return reparations

@api_router.get("/reparations/{reparation_id}", response_model=Reparation)
async def get_reparation(reparation_id: str):
    """Get a specific repair"""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    # Migrate old data
    if "description_panne" not in rep and "probleme_declare" in rep:
        rep["description_panne"] = rep.get("probleme_declare", "")
    if "tracking_id" not in rep:
        rep["tracking_id"] = str(uuid.uuid4())[:8].upper()
    if "heure_creation" not in rep:
        rep["heure_creation"] = ""
    if "statut_interne" not in rep:
        rep["statut_interne"] = "Terminé" if rep.get("statut") == "Terminé" else "En cours"
    
    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    if client:
        rep["client_nom"] = client.get("nom")
        rep["client_prenom"] = client.get("prenom")
        rep["client_email"] = client.get("email")
        rep["client_telephone"] = client.get("telephone")
    
    return rep

@api_router.put("/reparations/{reparation_id}", response_model=Reparation)
async def update_reparation(reparation_id: str, update: ReparationUpdate):
    """Update a repair"""
    existing = await db.reparations.find_one({"id": reparation_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["date_modification"] = datetime.now(timezone.utc).isoformat()

    # Synchronise le legacy "urgence" avec la présence de "urgence_89" dans forfaits
    if "forfaits" in update_data:
        update_data["urgence"] = "urgence_89" in (update_data["forfaits"] or [])

    await db.reparations.update_one({"id": reparation_id}, {"$set": update_data})
    
    updated = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    
    # Migrate old data
    if "description_panne" not in updated and "probleme_declare" in updated:
        updated["description_panne"] = updated.get("probleme_declare", "")
    if "tracking_id" not in updated:
        updated["tracking_id"] = str(uuid.uuid4())[:8].upper()
    if "heure_creation" not in updated:
        updated["heure_creation"] = ""
    if "statut_interne" not in updated:
        updated["statut_interne"] = "Terminé" if updated.get("statut") == "Terminé" else "En cours"
    
    client = await db.clients.find_one({"id": updated["client_id"]}, {"_id": 0})
    if client:
        updated["client_nom"] = client.get("nom")
        updated["client_prenom"] = client.get("prenom")
        updated["client_email"] = client.get("email")
        updated["client_telephone"] = client.get("telephone")
    
    return updated

@api_router.delete("/reparations/{reparation_id}")
async def delete_reparation(reparation_id: str):
    """Delete a repair"""
    result = await db.reparations.delete_one({"id": reparation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    return {"message": "Réparation supprimée"}

# ===================== SIGNATURE CLIENT =====================

class SignatureInput(BaseModel):
    signature_b64: str
    nom_signataire: Optional[str] = None
    accepte_conditions: bool = True

@api_router.get("/reparations/{reparation_id}/public")
async def get_reparation_public(reparation_id: str):
    """Fiche simplifiée pour le mode signature client (pas de données sensibles)"""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    materiel_list = get_materiel_fourni_list(rep.get("materiel_fourni", {}), rep.get("autre_materiel"))
    return {
        "id": rep["id"],
        "numero": rep.get("numero"),
        "date_creation": _fr_date(rep.get("date_creation")),
        "client_nom": client.get("nom") if client else "",
        "client_prenom": client.get("prenom") if client else "",
        "client_telephone": client.get("telephone") if client else "",
        "materiel": materiel_list,
        "description_panne": rep.get("description_panne", ""),
        "numero_serie": rep.get("numero_serie", ""),
        "etat_depot": rep.get("etat_depot", ""),
        "urgence": rep.get("urgence", False),
        "forfaits": rep.get("forfaits", []),
        "signature_b64": rep.get("signature_b64"),
        "date_signature": rep.get("date_signature"),
        "nom_signataire": rep.get("nom_signataire"),
        "conditions": CONDITIONS_REPARATION,
        "company": COMPANY_INFO,
    }

@api_router.post("/reparations/{reparation_id}/signature")
async def save_signature(reparation_id: str, payload: SignatureInput):
    """Enregistre la signature du client (écrase la précédente si présente)"""
    if not payload.accepte_conditions:
        raise HTTPException(status_code=400, detail="Les conditions de réparation doivent être acceptées")
    if not payload.signature_b64 or len(payload.signature_b64) < 100:
        raise HTTPException(status_code=400, detail="Signature invalide")

    rep = await db.reparations.find_one({"id": reparation_id})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")

    now = datetime.now(timezone.utc).isoformat()
    await db.reparations.update_one(
        {"id": reparation_id},
        {"$set": {
            "signature_b64": payload.signature_b64,
            "nom_signataire": payload.nom_signataire,
            "date_signature": now,
            "envoye_sans_signature": False,
            "date_modification": now,
        }}
    )
    # Libère l'iPad terminal si cette fiche y était assignée
    state = await db.ipad_state.find_one({"_id": "current"})
    if state and state.get("reparation_id") == reparation_id:
        await db.ipad_state.update_one(
            {"_id": "current"},
            {"$set": {"reparation_id": None, "assigned_at": None}},
        )
    return {"ok": True, "date_signature": now}

@api_router.delete("/reparations/{reparation_id}/signature")
async def delete_signature(reparation_id: str):
    """Supprime la signature (permet une re-signature)"""
    rep = await db.reparations.find_one({"id": reparation_id})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    await db.reparations.update_one(
        {"id": reparation_id},
        {"$set": {
            "signature_b64": None,
            "nom_signataire": None,
            "date_signature": None,
            "date_modification": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"ok": True}


class EncaisserReparationPayload(BaseModel):
    paiements: List[PaiementDetail]
    type_recette: Optional[str] = None  # par défaut déduit du prix
    remarque: Optional[str] = None


@api_router.post("/reparations/{reparation_id}/encaisser")
async def encaisser_reparation(reparation_id: str, payload: EncaisserReparationPayload):
    """Marque une réparation comme encaissée et crée automatiquement l'encaissement correspondant."""
    rep = await db.reparations.find_one({"id": reparation_id})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    if rep.get("encaissement_id"):
        raise HTTPException(status_code=400, detail="Cette réparation a déjà été encaissée")

    prix = float(rep.get("prix") or 0)
    if prix <= 0:
        raise HTTPException(status_code=400, detail="Veuillez saisir un prix sur la fiche avant d'encaisser")

    # Somme des paiements
    valid_paiements = [p for p in payload.paiements if p.mode and p.montant and p.montant > 0]
    if not valid_paiements:
        raise HTTPException(status_code=400, detail="Au moins un mode de paiement avec montant requis")
    total_pay = round(sum(p.montant for p in valid_paiements), 2)
    if abs(total_pay - prix) > 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"La somme des paiements ({total_pay:.2f} €) ne correspond pas au prix ({prix:.2f} €)"
        )

    # Type de recette : fourni OU déduit du prix
    type_recette = payload.type_recette
    if not type_recette:
        if abs(prix - 63) < 0.01: type_recette = "standard_63"
        elif abs(prix - 30) < 0.01: type_recette = "rapide_30"
        elif abs(prix - 10) < 0.01: type_recette = "express_10"
        elif abs(prix - 89) < 0.01: type_recette = "urgence_89"
        elif abs(prix - 45) < 0.01: type_recette = "imprimante_45"
        elif abs(prix - 79) < 0.01: type_recette = "recup_defectueux_79"
        elif abs(prix - 15) < 0.01: type_recette = "devis_15"
        else: type_recette = "autre"

    now = datetime.now(timezone.utc).isoformat()
    enc_id = str(uuid.uuid4())
    ht = round(prix / 1.2, 2)
    remarque = (payload.remarque or "").strip() or f"Encaissement fiche {rep.get('numero', '')}"

    enc_doc = {
        "id": enc_id,
        "type_recette": type_recette,
        "montant_ttc": prix,
        "montant_ht": ht,
        "paiements": [p.model_dump() for p in valid_paiements],
        "lignes": None,
        "client_id": rep.get("client_id"),
        "reparation_id": reparation_id,
        "reference": rep.get("numero"),
        "remarque": remarque,
        "date": now,
    }
    await db.encaissements.insert_one(enc_doc)
    enc_doc.pop("_id", None)

    # Met à jour la fiche réparation
    await db.reparations.update_one(
        {"id": reparation_id},
        {"$set": {
            "encaissement_id": enc_id,
            "date_paiement": now,
            "statut_interne": "Réglé",
            "statut": "Appareil prêt",
            "date_modification": now,
        }}
    )

    return {"ok": True, "encaissement": enc_doc}


@api_router.delete("/reparations/{reparation_id}/encaisser")
async def annuler_encaissement_reparation(reparation_id: str):
    """Annule l'encaissement d'une réparation (supprime l'encaissement lié)."""
    rep = await db.reparations.find_one({"id": reparation_id})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    enc_id = rep.get("encaissement_id")
    if not enc_id:
        raise HTTPException(status_code=400, detail="Cette réparation n'est pas encaissée")
    await db.encaissements.delete_one({"id": enc_id})
    await db.reparations.update_one(
        {"id": reparation_id},
        {"$set": {
            "encaissement_id": None,
            "date_paiement": None,
            "statut_interne": "En cours",
            "date_modification": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"ok": True}

# ===================== PUBLIC TRACKING =====================

@api_router.get("/suivi/{tracking_id}")
async def get_public_tracking(tracking_id: str):
    """Get public tracking info for a repair (no sensitive data)"""
    rep = await db.reparations.find_one({"tracking_id": tracking_id.upper()}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    
    # Return only public info
    materiel_list = get_materiel_fourni_list(
        rep.get('materiel_fourni', {}),
        rep.get('autre_materiel')
    )
    
    return {
        "numero": rep.get("numero"),
        "client_nom": client.get("nom") if client else "",
        "client_prenom": client.get("prenom") if client else "",
        "date_depot": rep.get("date_creation", "")[:10],
        "materiel": materiel_list,
        "statut": rep.get("statut"),
        "urgence": rep.get("urgence", False)
    }

# ===================== AUTH (JWT + bcrypt) =====================

JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

def _jwt_secret() -> str:
    secret = os.environ.get("JWT_SECRET")
    if not secret:
        raise RuntimeError("JWT_SECRET non configuré")
    return secret

def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def _verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def _create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS),
        "type": "access",
    }
    return pyjwt.encode(payload, _jwt_secret(), algorithm=JWT_ALGORITHM)

class LoginInput(BaseModel):
    email: str
    password: str

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else None
    if not token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    try:
        payload = pyjwt.decode(token, _jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token invalide")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur introuvable")
        return user
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Session expirée")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

@api_router.post("/auth/login")
async def login(payload: LoginInput, request: Request):
    email = payload.email.strip().lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"

    # Brute-force lockout check
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("locked_until"):
        try:
            locked_until = datetime.fromisoformat(attempt["locked_until"])
            if locked_until > datetime.now(timezone.utc):
                raise HTTPException(
                    status_code=429,
                    detail=f"Trop de tentatives. Réessayez dans {LOCKOUT_MINUTES} minutes.",
                )
        except (ValueError, TypeError):
            pass

    user = await db.users.find_one({"email": email})
    if not user or not _verify_password(payload.password, user.get("password_hash", "")):
        # Increment failed attempts
        new_count = (attempt.get("count", 0) if attempt else 0) + 1
        update = {"count": new_count, "last_attempt": datetime.now(timezone.utc).isoformat()}
        if new_count >= MAX_LOGIN_ATTEMPTS:
            update["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
        await db.login_attempts.update_one(
            {"identifier": identifier}, {"$set": {"identifier": identifier, **update}}, upsert=True,
        )
        raise HTTPException(status_code=401, detail="Identifiant ou mot de passe incorrect")

    # Success — reset attempts
    await db.login_attempts.delete_one({"identifier": identifier})

    token = _create_access_token(user["id"], user["email"])
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user["id"], "email": user["email"], "name": user.get("name", "")},
    }

@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return {"id": user["id"], "email": user["email"], "name": user.get("name", "")}

@api_router.post("/auth/logout")
async def logout(_: dict = Depends(get_current_user)):
    # Avec JWT stateless, le logout côté backend = no-op (le client supprime son token)
    return {"ok": True}

class ChangePasswordInput(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(payload: ChangePasswordInput, user: dict = Depends(get_current_user)):
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit contenir au moins 8 caractères")
    full = await db.users.find_one({"id": user["id"]})
    if not full or not _verify_password(payload.current_password, full.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Mot de passe actuel incorrect")
    if payload.current_password == payload.new_password:
        raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit être différent de l'actuel")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": _hash_password(payload.new_password), "password_updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"ok": True, "message": "Mot de passe mis à jour"}

# ===================== PRIVACY POLICY =====================

@api_router.get("/privacy-policy")
async def get_privacy_policy():
    return PRIVACY_POLICY

# ===================== IPAD TERMINAL STATE =====================

class IpadAssignInput(BaseModel):
    reparation_id: str
    kiosk: bool = True

ASSIGNMENT_TTL_SECONDS = 30 * 60  # 30 minutes

async def _get_ipad_state() -> dict:
    state = await db.ipad_state.find_one({"_id": "current"})
    if not state:
        state = {
            "_id": "current",
            "reparation_id": None,
            "assigned_at": None,
            "kiosk": True,
            "last_heartbeat_at": None,
        }
        await db.ipad_state.insert_one(state)
    return state

def _ipad_is_online(state: dict) -> bool:
    hb = state.get("last_heartbeat_at")
    if not hb:
        return False
    try:
        last = datetime.fromisoformat(hb.replace("Z", "+00:00")) if isinstance(hb, str) else hb
        now = datetime.now(timezone.utc)
        return (now - last).total_seconds() < 30
    except Exception:
        return False

@api_router.get("/ipad/current")
async def ipad_current():
    """iPad polling endpoint — renvoie la fiche assignée, ou null."""
    state = await _get_ipad_state()
    reparation_id = state.get("reparation_id")
    assigned_at = state.get("assigned_at")

    # Auto-expire stale assignments
    if reparation_id and assigned_at:
        try:
            ts = datetime.fromisoformat(assigned_at.replace("Z", "+00:00")) if isinstance(assigned_at, str) else assigned_at
            age = (datetime.now(timezone.utc) - ts).total_seconds()
            if age > ASSIGNMENT_TTL_SECONDS:
                await db.ipad_state.update_one(
                    {"_id": "current"},
                    {"$set": {"reparation_id": None, "assigned_at": None}},
                )
                reparation_id = None
                assigned_at = None
        except Exception:
            pass

    return {
        "reparation_id": reparation_id,
        "assigned_at": assigned_at,
        "kiosk": state.get("kiosk", True),
    }

@api_router.post("/ipad/assign")
async def ipad_assign(payload: IpadAssignInput):
    """PC → iPad : envoyer une fiche de réparation à signer."""
    rep = await db.reparations.find_one({"id": payload.reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    now = datetime.now(timezone.utc).isoformat()
    await db.ipad_state.update_one(
        {"_id": "current"},
        {"$set": {
            "reparation_id": payload.reparation_id,
            "assigned_at": now,
            "kiosk": bool(payload.kiosk),
        }},
        upsert=True,
    )
    return {"ok": True, "assigned_at": now, "kiosk": payload.kiosk}

@api_router.post("/ipad/release")
async def ipad_release():
    """Libère l'iPad (retour écran d'accueil)."""
    await db.ipad_state.update_one(
        {"_id": "current"},
        {"$set": {"reparation_id": None, "assigned_at": None}},
        upsert=True,
    )
    return {"ok": True}

@api_router.put("/ipad/heartbeat")
async def ipad_heartbeat():
    """L'iPad signale qu'il est connecté (toutes les 3-10 s)."""
    now = datetime.now(timezone.utc).isoformat()
    await db.ipad_state.update_one(
        {"_id": "current"},
        {"$set": {"last_heartbeat_at": now}},
        upsert=True,
    )
    return {"ok": True, "at": now}

@api_router.get("/ipad/status")
async def ipad_status():
    """PC → indicateur 'iPad en ligne' + infos assignation courante."""
    state = await _get_ipad_state()
    return {
        "online": _ipad_is_online(state),
        "last_heartbeat_at": state.get("last_heartbeat_at"),
        "reparation_id": state.get("reparation_id"),
        "assigned_at": state.get("assigned_at"),
        "kiosk": state.get("kiosk", True),
    }

# ===================== PDF GENERATION =====================

@api_router.get("/reparations/{reparation_id}/pdf/client")
async def get_client_pdf(reparation_id: str):
    """Generate and return client PDF"""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Generate tracking URL
    frontend_url = os.environ.get('FRONTEND_URL', 'https://fiche-repair.preview.emergentagent.com')
    tracking_url = f"{frontend_url}/suivi/{rep.get('tracking_id', '')}"
    
    pdf_content = generate_client_pdf(rep, client, tracking_url)
    
    filename = sanitize_filename(f"Reparation_n{rep['numero']}_client.pdf")
    filepath = PDF_DIR / filename
    with open(filepath, 'wb') as f:
        f.write(pdf_content)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'}
    )

@api_router.get("/reparations/{reparation_id}/pdf/interne")
async def get_internal_pdf(reparation_id: str):
    """Generate and return internal PDF"""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    pdf_content = generate_internal_pdf(rep, client)
    
    filename = sanitize_filename(f"Reparation_n{rep['numero']}_interne.pdf")
    filepath = PDF_DIR / filename
    with open(filepath, 'wb') as f:
        f.write(pdf_content)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'}
    )


# ---------- Paramètres : bannière publicitaire du compte-rendu ----------

class AdBannerPayload(BaseModel):
    image_b64: str  # data URI ou base64 brut (JPG/PNG)


async def _get_ad_banner_bytes() -> Optional[bytes]:
    doc = await db.settings.find_one({"key": "ad_banner"}, {"_id": 0})
    if not doc or not doc.get("image_b64"):
        return None
    raw = doc["image_b64"]
    if raw.startswith("data:"):
        raw = raw.split(",", 1)[1]
    try:
        return base64.b64decode(raw)
    except Exception:
        return None


@api_router.get("/settings/ad-banner")
async def get_ad_banner():
    """Retourne la bannière publicitaire stockée (métadonnées + image base64)."""
    doc = await db.settings.find_one({"key": "ad_banner"}, {"_id": 0})
    if not doc:
        return {"exists": False, "image_b64": None, "updated_at": None}
    return {
        "exists": bool(doc.get("image_b64")),
        "image_b64": doc.get("image_b64"),
        "updated_at": doc.get("updated_at"),
    }


@api_router.put("/settings/ad-banner")
async def put_ad_banner(payload: AdBannerPayload, user: dict = Depends(get_current_user)):
    """Upload/remplace la bannière publicitaire. Limite ~3 Mo."""
    raw = payload.image_b64 or ""
    if len(raw) > 4_500_000:  # ~3.3 Mo après base64
        raise HTTPException(status_code=413, detail="Image trop lourde (limite ~3 Mo)")
    if len(raw) < 100:
        raise HTTPException(status_code=400, detail="Image invalide")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.settings.update_one(
        {"key": "ad_banner"},
        {"$set": {"key": "ad_banner", "image_b64": raw, "updated_at": now_iso}},
        upsert=True,
    )
    return {"success": True, "updated_at": now_iso}


@api_router.delete("/settings/ad-banner")
async def delete_ad_banner(user: dict = Depends(get_current_user)):
    await db.settings.delete_one({"key": "ad_banner"})
    return {"success": True}


@api_router.get("/reparations/{reparation_id}/pdf/compte-rendu")
async def get_compte_rendu_pdf(reparation_id: str):
    """Fiche compte rendu à remettre au client après la réparation."""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")

    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")

    ad_bytes = await _get_ad_banner_bytes()
    pdf_content = generate_compte_rendu_pdf(rep, client, ad_banner_bytes=ad_bytes)

    filename = sanitize_filename(f"Reparation_n{rep['numero']}_compte_rendu.pdf")
    filepath = PDF_DIR / filename
    with open(filepath, 'wb') as f:
        f.write(pdf_content)

    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'}
    )


@api_router.get("/reparations/{reparation_id}/qrcode")
async def get_qrcode(reparation_id: str):
    """Generate QR code for repair tracking"""
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://fiche-repair.preview.emergentagent.com')
    tracking_url = f"{frontend_url}/suivi/{rep.get('tracking_id', '')}"
    
    qr_bytes = generate_qr_code(tracking_url)
    
    return Response(
        content=qr_bytes,
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="qr_{rep["numero"]}.png"'}
    )

# ===================== EMAIL =====================

@api_router.post("/reparations/{reparation_id}/send-email")
async def send_repair_email(reparation_id: str, force: bool = Query(False, description="Forcer l'envoi sans signature")):
    """Send repair email to client with PDF attachment. Blocks if no signature unless force=true."""
    # First, check reparation and signature status BEFORE checking Resend config
    rep = await db.reparations.find_one({"id": reparation_id}, {"_id": 0})
    if not rep:
        raise HTTPException(status_code=404, detail="Réparation non trouvée")

    client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")

    if not client.get("email"):
        raise HTTPException(status_code=400, detail="Le client n'a pas d'adresse email")

    # Check signature BEFORE checking Resend API key (per requirements)
    has_signature = bool(rep.get("signature_b64"))
    if not has_signature and not force:
        raise HTTPException(
            status_code=409,
            detail="Impossible d'envoyer : le client n'a pas signé. Utilisez force=true pour envoyer sans signature.",
        )

    # Marquer l'envoi forcé sans signature (do this BEFORE checking Resend)
    if not has_signature and force:
        await db.reparations.update_one(
            {"id": reparation_id},
            {"$set": {"envoye_sans_signature": True, "date_modification": datetime.now(timezone.utc).isoformat()}},
        )
        rep["envoye_sans_signature"] = True

    # Now check Resend configuration
    if not resend.api_key:
        raise HTTPException(status_code=500, detail="Service email non configuré")

    frontend_url = os.environ.get('FRONTEND_URL', 'https://fiche-repair.preview.emergentagent.com')
    tracking_url = f"{frontend_url}/suivi/{rep.get('tracking_id', '')}"
    
    pdf_content = generate_client_pdf(rep, client, tracking_url)
    pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
    filename = sanitize_filename(f"Reparation_n{rep['numero']}.pdf")
    
    html_content = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333;">
        <p>Bonjour {client.get('prenom', '')} {client.get('nom', '')},</p>
        <p>Veuillez trouver en pièce jointe votre fiche de réparation n°{rep['numero']}.</p>
        <p>Vous pouvez suivre l'avancement de votre réparation en ligne :</p>
        <p><a href="{tracking_url}" style="color: #84CC16; font-weight: bold;">{tracking_url}</a></p>
        <p>Cordialement,<br>
        <strong>{COMPANY_INFO['name']}</strong><br>
        {COMPANY_INFO['address']}<br>
        Tél: {COMPANY_INFO['phone']}<br>
        Email: {COMPANY_INFO['email']}</p>
    </body>
    </html>
    """
    
    params = {
        "from": SENDER_EMAIL,
        "to": [client["email"]],
        "subject": f"Fiche de réparation n°{rep['numero']} - {COMPANY_INFO['name']}",
        "html": html_content,
        "attachments": [{"filename": filename, "content": pdf_base64}]
    }
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, params)
        return {"status": "success", "message": f"Email envoyé à {client['email']}", "email_id": email.get("id")}
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'envoi: {str(e)}")

# ===================== COMMANDES CRUD =====================

@api_router.get("/commandes/statuts")
async def get_statuts_commande():
    """Get available order statuses"""
    return {"statuts": STATUTS_COMMANDE}

@api_router.post("/commandes", response_model=Commande)
async def create_commande(commande: CommandeCreate):
    """Create a new order"""
    client = await db.clients.find_one({"id": commande.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    now = datetime.now(timezone.utc).isoformat()
    numero = await get_next_commande_number()
    
    montant_total = None
    if commande.prix_vente and commande.quantite:
        montant_total = commande.prix_vente * commande.quantite
    
    cmd_doc = {
        "id": str(uuid.uuid4()),
        "numero": numero,
        "client_id": commande.client_id,
        "reference_produit": commande.reference_produit,
        "designation": commande.designation,
        "fournisseur": commande.fournisseur,
        "quantite": commande.quantite,
        "prix_achat": commande.prix_achat,
        "prix_vente": commande.prix_vente,
        "montant_total": montant_total,
        "statut": commande.statut,
        "remarques": commande.remarques,
        "date_creation": now,
        "date_modification": now
    }
    
    await db.commandes.insert_one(cmd_doc)
    cmd_doc.pop("_id", None)
    
    cmd_doc["client_nom"] = client.get("nom")
    cmd_doc["client_prenom"] = client.get("prenom")
    cmd_doc["client_telephone"] = client.get("telephone")
    
    return cmd_doc

@api_router.get("/commandes", response_model=List[Commande])
async def get_commandes(
    search: Optional[str] = Query(None),
    statut: Optional[str] = Query(None),
    limit: int = Query(100, le=500)
):
    """Get all orders with optional filters"""
    query = {}
    if statut:
        query["statut"] = statut
    
    commandes = await db.commandes.find(query, {"_id": 0}).sort("date_creation", -1).to_list(limit)
    
    for cmd in commandes:
        client = await db.clients.find_one({"id": cmd["client_id"]}, {"_id": 0})
        if client:
            cmd["client_nom"] = client.get("nom")
            cmd["client_prenom"] = client.get("prenom")
            cmd["client_telephone"] = client.get("telephone")
    
    if search:
        search_normalized = unidecode(search.lower())
        filtered = []
        for cmd in commandes:
            client_name = f"{cmd.get('client_prenom', '')} {cmd.get('client_nom', '')}".lower()
            client_name_normalized = unidecode(client_name)
            numero = cmd.get("numero", "").lower()
            designation = unidecode(cmd.get("designation", "").lower())
            
            if (search_normalized in client_name_normalized or
                search_normalized in numero or
                search_normalized in designation):
                filtered.append(cmd)
        return filtered
    
    return commandes

@api_router.get("/commandes/{commande_id}", response_model=Commande)
async def get_commande(commande_id: str):
    """Get a specific order"""
    cmd = await db.commandes.find_one({"id": commande_id}, {"_id": 0})
    if not cmd:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    client = await db.clients.find_one({"id": cmd["client_id"]}, {"_id": 0})
    if client:
        cmd["client_nom"] = client.get("nom")
        cmd["client_prenom"] = client.get("prenom")
        cmd["client_telephone"] = client.get("telephone")
    
    return cmd

@api_router.put("/commandes/{commande_id}", response_model=Commande)
async def update_commande(commande_id: str, update: CommandeUpdate):
    """Update an order"""
    existing = await db.commandes.find_one({"id": commande_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["date_modification"] = datetime.now(timezone.utc).isoformat()
    
    # Recalculate total
    prix_vente = update_data.get("prix_vente", existing.get("prix_vente"))
    quantite = update_data.get("quantite", existing.get("quantite"))
    if prix_vente and quantite:
        update_data["montant_total"] = prix_vente * quantite
    
    await db.commandes.update_one({"id": commande_id}, {"$set": update_data})
    
    updated = await db.commandes.find_one({"id": commande_id}, {"_id": 0})
    
    client = await db.clients.find_one({"id": updated["client_id"]}, {"_id": 0})
    if client:
        updated["client_nom"] = client.get("nom")
        updated["client_prenom"] = client.get("prenom")
        updated["client_telephone"] = client.get("telephone")
    
    return updated

@api_router.delete("/commandes/{commande_id}")
async def delete_commande(commande_id: str):
    """Delete an order"""
    result = await db.commandes.delete_one({"id": commande_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    return {"message": "Commande supprimée"}

@api_router.delete("/commandes/purge/completed")
async def purge_completed_commandes():
    """Delete all completed orders (Livré/Récupéré, Réglé)"""
    result = await db.commandes.delete_many({
        "statut": {"$in": ["Livré/Récupéré", "Réglé"]}
    })
    return {"message": f"{result.deleted_count} commandes supprimées"}

# ===================== ENCAISSEMENT =====================

@api_router.get("/encaissements/types")
async def get_types_recette():
    """Get available receipt types"""
    return {"types": TYPES_RECETTE}

@api_router.post("/encaissements", response_model=Encaissement)
async def create_encaissement(encaissement: EncaissementCreate):
    """Create a new receipt entry with multiple payment methods and optional multi-line detail."""
    now = datetime.now(timezone.utc).isoformat()

    # Lignes : si fournies, on calcule / valide les totaux depuis les lignes
    lignes_dump = None
    type_recette = encaissement.type_recette
    montant_ttc = encaissement.montant_ttc
    montant_ht = encaissement.montant_ht

    if encaissement.lignes and len(encaissement.lignes) > 0:
        lignes_dump = [l.model_dump() for l in encaissement.lignes]
        sum_ttc = round(sum(l["montant_ttc"] for l in lignes_dump), 2)
        # La somme des lignes doit correspondre au montant TTC global
        if abs(sum_ttc - montant_ttc) > 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"La somme des lignes ({sum_ttc:.2f} €) ne correspond pas au total TTC ({montant_ttc:.2f} €)"
            )
        # Si plusieurs lignes et type non spécifique → marquer "mixte"
        if len(lignes_dump) > 1 and type_recette not in (None, "", "mixte"):
            # Si une seule catégorie utilisée, on garde ; sinon on force "mixte"
            categories = {l["type_recette"] for l in lignes_dump}
            if len(categories) > 1:
                type_recette = "mixte"
        elif len(lignes_dump) == 1:
            type_recette = lignes_dump[0]["type_recette"]

    # Calculate HT if not provided (assuming 20% TVA)
    if not montant_ht:
        montant_ht = round(montant_ttc / 1.2, 2)

    entry_doc = {
        "id": str(uuid.uuid4()),
        "type_recette": type_recette,
        "montant_ttc": montant_ttc,
        "montant_ht": montant_ht,
        "paiements": [p.model_dump() for p in encaissement.paiements],
        "lignes": lignes_dump,
        "client_id": encaissement.client_id,
        "reparation_id": encaissement.reparation_id,
        "reference": encaissement.reference,
        "remarque": encaissement.remarque,
        "date": now
    }

    await db.encaissements.insert_one(entry_doc)
    entry_doc.pop("_id", None)

    if encaissement.client_id:
        client = await db.clients.find_one({"id": encaissement.client_id}, {"_id": 0})
        if client:
            entry_doc["client_nom"] = client.get("nom")
            entry_doc["client_prenom"] = client.get("prenom")

    return entry_doc

@api_router.get("/encaissements")
async def get_encaissements(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = Query(100, le=500)
):
    """Get receipt entries"""
    query = {}
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        if date_query:
            query["date"] = date_query
    
    entries = await db.encaissements.find(query, {"_id": 0}).sort("date", -1).to_list(limit)
    
    for entry in entries:
        # Migration for old entries
        if "montant" in entry and "montant_ttc" not in entry:
            entry["montant_ttc"] = entry["montant"]
            entry["montant_ht"] = round(entry["montant"] / 1.2, 2)
        if "mode_paiement" in entry and "paiements" not in entry:
            entry["paiements"] = [{"mode": entry["mode_paiement"], "montant": entry.get("montant_ttc", entry.get("montant", 0))}]
        
        if entry.get("client_id"):
            client = await db.clients.find_one({"id": entry["client_id"]}, {"_id": 0})
            if client:
                entry["client_nom"] = client.get("nom")
                entry["client_prenom"] = client.get("prenom")
    
    return entries

@api_router.delete("/encaissements/{entry_id}")
async def delete_encaissement(entry_id: str):
    """Delete a receipt entry"""
    result = await db.encaissements.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entrée non trouvée")
    return {"message": "Entrée supprimée"}

# ===================== CAISSE (JOURNAL COMPLET) =====================

@api_router.post("/caisse", response_model=CaisseEntry)
async def create_caisse_entry(entry: CaisseEntryCreate):
    """Create a new cash register entry"""
    now = datetime.now(timezone.utc).isoformat()
    
    entry_doc = {
        "id": str(uuid.uuid4()),
        "type": entry.type,
        "montant": entry.montant,
        "description": entry.description,
        "mode_paiement": entry.mode_paiement,
        "reparation_id": entry.reparation_id,
        "client_id": entry.client_id,
        "date": now
    }
    
    await db.caisse.insert_one(entry_doc)
    entry_doc.pop("_id", None)
    return entry_doc

@api_router.get("/caisse", response_model=List[CaisseEntry])
async def get_caisse_entries(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = Query(500, le=2000)
):
    """Get cash register entries — unified view (caisse + encaissements)"""
    query = {}
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        if date_query:
            query["date"] = date_query

    # Manual caisse entries
    entries = await db.caisse.find(query, {"_id": 0}).to_list(limit)

    # Also include encaissements (recettes quotidiennes) comme entrées de caisse automatiques
    encaissements = await db.encaissements.find(query, {"_id": 0}).to_list(limit)

    type_labels = {
        "forfait_63": "Forfait réparation 63€",
        "rapide_30": "Réparation rapide 30€",
        "express_10": "Réparation express 10€",
        "devis_15": "Devis 15€",
        "ventes": "Ventes",
        "autre": "Autre recette",
    }

    for enc in encaissements:
        total_ttc = enc.get("montant_ttc", enc.get("montant", 0)) or 0
        paiements = enc.get("paiements") or ([{"mode": enc.get("mode_paiement"), "montant": total_ttc}] if enc.get("mode_paiement") else [])
        # Créer une ligne caisse par mode de paiement (vue unifiée)
        type_label = type_labels.get(enc.get("type_recette"), enc.get("type_recette", "Recette"))
        remarque = enc.get("remarque") or ""
        description_base = f"{type_label}" + (f" — {remarque}" if remarque else "")
        if paiements:
            for p in paiements:
                entries.append({
                    "id": f"enc-{enc['id']}-{p.get('mode','')}",
                    "type": "entree",
                    "montant": float(p.get("montant") or 0),
                    "description": description_base,
                    "mode_paiement": p.get("mode"),
                    "reparation_id": None,
                    "client_id": enc.get("client_id"),
                    "date": enc.get("date"),
                })
        else:
            entries.append({
                "id": f"enc-{enc['id']}",
                "type": "entree",
                "montant": float(total_ttc),
                "description": description_base,
                "mode_paiement": None,
                "reparation_id": None,
                "client_id": enc.get("client_id"),
                "date": enc.get("date"),
            })

    # Tri décroissant par date
    entries.sort(key=lambda e: e.get("date") or "", reverse=True)
    return entries[:limit]

@api_router.delete("/caisse/{entry_id}")
async def delete_caisse_entry(entry_id: str):
    """Delete a cash register entry (or underlying encaissement if id starts with 'enc-')"""
    if entry_id.startswith("enc-"):
        # id format: enc-{encaissement_id} ou enc-{encaissement_id}-{mode}
        parts = entry_id.split("-")
        # UUID4 has 5 parts separated by '-', so we rebuild it
        if len(parts) >= 6:
            enc_id = "-".join(parts[1:6])
        else:
            enc_id = "-".join(parts[1:])
        result = await db.encaissements.delete_one({"id": enc_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Encaissement non trouvé")
        return {"message": "Encaissement supprimé"}

    result = await db.caisse.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entrée non trouvée")
    return {"message": "Entrée supprimée"}

# ===================== EXPORTS =====================

@api_router.get("/export/reparations/excel")
async def export_reparations_excel(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None)
):
    """Export repairs to Excel"""
    query = {}
    if date_from:
        query["date_creation"] = {"$gte": date_from}
    if date_to:
        if "date_creation" in query:
            query["date_creation"]["$lte"] = date_to + "T23:59:59"
        else:
            query["date_creation"] = {"$lte": date_to + "T23:59:59"}
    
    reparations = await db.reparations.find(query, {"_id": 0}).sort("date_creation", -1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Réparations"
    
    header_fill = PatternFill(start_color="84CC16", end_color="84CC16", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["N°", "Date", "Client", "Téléphone", "Matériel", "Description", "Diagnostic", "Action", "Prix", "Statut", "Urgence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, rep in enumerate(reparations, 2):
        client = await db.clients.find_one({"id": rep["client_id"]}, {"_id": 0})
        client_name = f"{client.get('prenom', '')} {client.get('nom', '')}" if client else "-"
        client_phone = client.get("telephone", "-") if client else "-"
        materiel_list = get_materiel_fourni_list(rep.get('materiel_fourni', {}), rep.get('autre_materiel'))
        
        data = [
            rep.get("numero", ""),
            rep.get("date_creation", "")[:10],
            client_name,
            client_phone,
            ", ".join(materiel_list),
            rep.get("description_panne", ""),
            rep.get("diagnostic", "") or "",
            rep.get("action_realisee", "") or "",
            rep.get("prix", "") or "",
            rep.get("statut", ""),
            "Oui" if rep.get("urgence") else "Non"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"reparations_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@api_router.get("/export/caisse/excel")
async def export_caisse_excel(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    include_empty_months: bool = Query(True, description="Create a tab for every month between first and last entry, even empty ones"),
):
    """
    Export Journal de caisse Excel — UN SEUL FICHIER avec UN ONGLET PAR MOIS.
    Structure identique au modèle DCLIC : colonnes A-S, formules Excel actives
    (TOTAL CA, SOLDE CAISSE, report M-1, TOTAL MOIS, HT/TVA) + onglet TOTAUX.
    """
    # ---------- 1. Récupération des données ----------
    query = {}
    if date_from or date_to:
        d = {}
        if date_from:
            d["$gte"] = date_from
        if date_to:
            d["$lte"] = date_to + "T23:59:59"
        query["date"] = d
    elif year and month:
        start = f"{year}-{month:02d}-01"
        end = f"{year+1}-01-01" if month == 12 else f"{year}-{month+1:02d}-01"
        query["date"] = {"$gte": start, "$lt": end}

    caisse_entries = await db.caisse.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    encaissement_entries = await db.encaissements.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    TYPE_LABELS = {
        "forfait_63": "Forfait réparation 63€",
        "rapide_30": "Réparation rapide 30€",
        "express_10": "Réparation express 10€",
        "devis_15": "Devis 15€",
        "ventes": "Ventes",
        "autre": "Autre recette",
    }

    # ---------- 2. Agrégation par (mois -> date -> totaux) ----------
    # {month_key: {date_str: {especes, cheques, cb, virement, depenses, remarques, reglement, numero, nom}}}
    months_data: Dict[str, Dict[str, Dict[str, Any]]] = {}

    def ensure_day(month_key: str, d: str):
        months_data.setdefault(month_key, {})
        months_data[month_key].setdefault(d, {
            "especes": 0.0, "cheques": 0.0, "cb": 0.0, "virement": 0.0,
            "depenses": 0.0, "remarques": [], "reglement": [],
            "numero": [], "nom": [],
        })
        return months_data[month_key][d]

    for e in caisse_entries:
        date_iso = e.get("date", "")
        if not date_iso:
            continue
        mk = date_iso[:7]
        d = date_iso[:10]
        row = ensure_day(mk, d)
        amount = float(e.get("montant") or 0)
        mode = (e.get("mode_paiement") or "").lower()
        if e.get("type") == "entree":
            if mode == "especes":
                row["especes"] += amount
            elif mode == "cheque":
                row["cheques"] += amount
            elif mode == "cb":
                row["cb"] += amount
            elif mode == "virement":
                row["virement"] += amount
        else:  # sortie
            row["depenses"] += amount
        if e.get("description"):
            row["remarques"].append(e["description"])

    for enc in encaissement_entries:
        date_iso = enc.get("date", "")
        if not date_iso:
            continue
        mk = date_iso[:7]
        d = date_iso[:10]
        row = ensure_day(mk, d)
        paiements = enc.get("paiements")
        total_ttc = float(enc.get("montant_ttc", enc.get("montant", 0)) or 0)
        if isinstance(paiements, list) and paiements:
            for p in paiements:
                mode = (p.get("mode") or "").lower()
                montant = float(p.get("montant") or 0)
                if mode == "especes":
                    row["especes"] += montant
                elif mode == "cheque":
                    row["cheques"] += montant
                elif mode == "cb":
                    row["cb"] += montant
                elif mode == "virement":
                    row["virement"] += montant
        else:
            mode = (enc.get("mode_paiement") or "").lower()
            if mode == "especes":
                row["especes"] += total_ttc
            elif mode == "cheque":
                row["cheques"] += total_ttc
            elif mode == "cb":
                row["cb"] += total_ttc
            elif mode == "virement":
                row["virement"] += total_ttc
        # Libellé : type + remarque
        label = TYPE_LABELS.get(enc.get("type_recette"), enc.get("type_recette") or "Recette")
        if enc.get("remarque"):
            label = f"{label} — {enc['remarque']}"
        row["remarques"].append(label)
        if enc.get("reference"):
            row["numero"].append(enc["reference"])
        if enc.get("client_prenom") or enc.get("client_nom"):
            row["nom"].append(f"{enc.get('client_prenom','')} {enc.get('client_nom','')}".strip())

    # ---------- 3. Si include_empty_months : créer une clé vide pour chaque mois entre min et max ----------
    if months_data and include_empty_months:
        from calendar import monthrange
        keys = sorted(months_data.keys())
        y, m = map(int, keys[0].split("-"))
        y2, m2 = map(int, keys[-1].split("-"))
        while (y, m) <= (y2, m2):
            months_data.setdefault(f"{y}-{m:02d}", {})
            m += 1
            if m > 12:
                m = 1
                y += 1

    # ---------- 4. Création du classeur ----------
    wb = Workbook()
    wb.remove(wb.active)  # on crée tous les onglets manuellement

    MONTH_ABBR = {
        1: "Jan", 2: "Fév", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
        7: "Juillet", 8: "Aout", 9: "Sept", 10: "Oct", 11: "Nov", 12: "Déc",
    }

    def sheet_name_for(month_key: str) -> str:
        y, m = map(int, month_key.split("-"))
        return f"{MONTH_ABBR[m]} {str(y)[-2:]}"

    header_fill = PatternFill(start_color="84CC16", end_color="84CC16", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sub_font = Font(italic=True, size=9, color="475569")
    bold_font = Font(bold=True, size=10)
    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fonds distinctifs (demande user msg 181) :
    # - Vert pâle  = cellule remplie automatiquement (calculée ou importée depuis Encaissement/Caisse)
    # - Bleu pâle  = cellule à remplir manuellement par le gérant
    # - Gris clair = colonnes "Règlement" (O, P, Q)
    # - Beige      = colonnes "Facturation externe" (R, S, T, U)
    auto_fill = PatternFill(start_color="ECFCCB", end_color="ECFCCB", fill_type="solid")   # vert pâle
    manual_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # bleu pâle
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # gris très clair
    gray_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")   # gris
    beige_fill = PatternFill(start_color="F5E6CA", end_color="F5E6CA", fill_type="solid")  # beige

    # Header (ligne 1 & 2) : vert DCLIC par défaut, sauf colonnes spécifiques
    GRAY_COLS = {15, 16, 17}          # O, P, Q
    BEIGE_COLS = {18, 19, 20, 21}     # R, S, T, U

    # Colonnes calculées/auto-remplies (A..U)
    AUTO_COLS = {1, 2, 3, 4, 6, 9, 12, 13}  # A,B,C,D,F,I,L,M
    # Colonnes à remplir manuellement (hors gray/beige)
    MANUAL_COLS = {5, 7, 8, 10, 11, 14}
    # E,G,H,J,K,N

    # Colonnes de la feuille (A..U)
    HEADERS_ROW1 = [
        "DATE", "ESPECES", "CHEQUES", "CB", "PNF", "TOTAL CA",
        "ENCAISSEM", "DEPENSES", "SOLDE CAISSE",
        "CLIENTS FACTURéS", "",  # J:K fusionnées
        "TOTAL ", "TOTAL",        # L:M fusionnées
        "REMARQUES",
        "Virements",              # O1
        "N° facture",             # P1
        "NOM\nFACTURE",
        "GoCardLess\nmontant net", "Frais de\ncommission",
        "Nom", "Facture/échéancier",  # T, U
    ]
    HEADERS_ROW2 = [
        "REPORT M-1", "", "", "", "", "",
        "clients", "fourniss", "",     # I2 sera rempli par formule
        "chèques", "CB",
        "REM CHQ", "REM CB",
        "",
        "",   # O2 vide
        "",   # P2 vide
        "",
        "", "",
        "", "",  # T2, U2
    ]
    COL_WIDTHS = {
        "A": 22, "B": 10, "C": 10, "D": 10, "E": 8, "F": 12,
        "G": 12, "H": 12, "I": 14,
        "J": 10, "K": 10, "L": 10, "M": 10,
        "N": 24, "O": 12, "P": 14, "Q": 20, "R": 12, "S": 12,
        "T": 18, "U": 22,
    }

    ordered_month_keys = sorted(months_data.keys())
    month_totals_by_sheet: List[Tuple[str, str, int]] = []  # [(month_key, sheet_name, total_mois_row_idx)]

    for idx, month_key in enumerate(ordered_month_keys):
        sname = sheet_name_for(month_key)[:31]
        ws = wb.create_sheet(title=sname)

        # Row 1 headers
        for col_i, h in enumerate(HEADERS_ROW1, start=1):
            c = ws.cell(row=1, column=col_i, value=h)
            if col_i in GRAY_COLS:
                c.fill = gray_fill
                c.font = Font(bold=True, color="111827", size=10)
            elif col_i in BEIGE_COLS:
                c.fill = beige_fill
                c.font = Font(bold=True, color="111827", size=10)
            else:
                c.fill = header_fill
                c.font = header_font
            c.alignment = center
            c.border = border
        # Fusions ligne 1
        ws.merge_cells("J1:K1")
        ws.merge_cells("L1:M1")

        # Row 2 sub-headers
        for col_i, h in enumerate(HEADERS_ROW2, start=1):
            c = ws.cell(row=2, column=col_i, value=h)
            if col_i in GRAY_COLS:
                c.fill = gray_fill
            elif col_i in BEIGE_COLS:
                c.fill = beige_fill
            c.font = sub_font
            c.alignment = center
            c.border = border
        # I2 = report du solde caisse du mois précédent — format monétaire, aligné à droite
        if idx == 0:
            i2_cell = ws.cell(row=2, column=9, value=0)
        else:
            prev_month_key, prev_sname, prev_total_row = month_totals_by_sheet[idx - 1]
            i2_cell = ws.cell(row=2, column=9, value=f"='{prev_sname}'!I{prev_total_row}")
        i2_cell.number_format = '#,##0.00 €'
        i2_cell.alignment = Alignment(horizontal="right", vertical="center")
        i2_cell.font = bold_font

        # Rows 3.. : une ligne par jour trié
        days = sorted(months_data[month_key].keys())
        last_data_row = 2
        for i, d in enumerate(days):
            r = 3 + i
            day = months_data[month_key][d]
            # A: DATE au format "jeudi 4 janvier" (locale FR)
            try:
                cell_a = ws.cell(row=r, column=1, value=datetime.fromisoformat(d))
                cell_a.number_format = '[$-40C]dddd d mmmm'
            except Exception:
                ws.cell(row=r, column=1, value=d)
            # B/C/D: ESPECES/CHEQUES/CB
            if day["especes"]:
                ws.cell(row=r, column=2, value=round(day["especes"], 2))
            if day["cheques"]:
                ws.cell(row=r, column=3, value=round(day["cheques"], 2))
            if day["cb"]:
                ws.cell(row=r, column=4, value=round(day["cb"], 2))
            # F: TOTAL CA = SUM(B:D)
            ws.cell(row=r, column=6, value=f"=SUM(B{r}:D{r})")
            # H: DEPENSES
            if day["depenses"]:
                ws.cell(row=r, column=8, value=round(day["depenses"], 2))
            # I: SOLDE CAISSE = I_prev + B + G - H
            ws.cell(row=r, column=9, value=f"=I{r-1}+B{r}+G{r}-H{r}")
            # L: REM CHQ = C + J  ; M: REM CB = D + K
            ws.cell(row=r, column=12, value=f"=C{r}+J{r}")
            ws.cell(row=r, column=13, value=f"=D{r}+K{r}")
            # N: REMARQUES — laissé VIDE (demande user msg 181 : à remplir à la main)
            # O: REGLEMENT (virements)
            if day["virement"]:
                ws.cell(row=r, column=15, value=round(day["virement"], 2))
            # P: NUMERO
            # P: N° facture — laissé VIDE (colonne manuelle, demande user)
            # Q: NOM FACTURE
            if day["nom"]:
                ws.cell(row=r, column=17, value=", ".join(day["nom"]))

            # Bordures + fond de couleur sur TOUTES les cellules de la ligne (A..U)
            for col_i in range(1, 22):
                cell = ws.cell(row=r, column=col_i)
                cell.border = border
                if col_i in GRAY_COLS:
                    cell.fill = gray_fill
                elif col_i in BEIGE_COLS:
                    cell.fill = beige_fill
                elif col_i in AUTO_COLS:
                    cell.fill = auto_fill
                elif col_i in MANUAL_COLS:
                    cell.fill = manual_fill
                if col_i in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 18, 19):
                    cell.number_format = "#,##0.00"
            last_data_row = r

        # Ligne TOTAL MOIS
        total_row = max(last_data_row + 1, 3)
        ws.cell(row=total_row, column=1, value="TOTAL MOIS").font = bold_font
        for col_letter, col_i in [("B", 2), ("C", 3), ("D", 4), ("E", 5), ("F", 6),
                                   ("G", 7), ("H", 8), ("J", 10), ("K", 11),
                                   ("L", 12), ("M", 13), ("O", 15), ("R", 18), ("S", 19)]:
            if last_data_row >= 3:
                cell = ws.cell(row=total_row, column=col_i,
                               value=f"=SUM({col_letter}3:{col_letter}{last_data_row})")
                cell.font = bold_font
                cell.number_format = "#,##0.00"
        # I TOTAL MOIS = solde caisse final
        if last_data_row >= 3:
            ws.cell(row=total_row, column=9, value=f"=I{last_data_row}").font = bold_font
        else:
            ws.cell(row=total_row, column=9, value=f"=I2").font = bold_font
        # Bordures + fond gris clair sur toute la ligne TOTAL MOIS
        for col_i in range(1, 22):
            cell = ws.cell(row=total_row, column=col_i)
            cell.border = border
            cell.fill = total_fill
            if not cell.font or not cell.font.bold:
                cell.font = bold_font

        # Ligne HT / TVA
        ht_row = total_row + 2
        tva_row = total_row + 3
        ws.cell(row=ht_row, column=5, value="HT").font = bold_font
        c = ws.cell(row=ht_row, column=6, value=f"=F{total_row}/1.2")
        c.font = bold_font
        c.number_format = "#,##0.00"
        ws.cell(row=tva_row, column=5, value="TVA").font = bold_font
        c = ws.cell(row=tva_row, column=6, value=f"=F{ht_row}*0.2")
        c.font = bold_font
        c.number_format = "#,##0.00"
        for rr in (ht_row, tva_row):
            for col_i in (5, 6):
                cell = ws.cell(row=rr, column=col_i)
                cell.border = border
                cell.fill = total_fill

        # Largeurs colonnes
        for col_letter, w in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = w
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A3"

        month_totals_by_sheet.append((month_key, sname, total_row))

    # ---------- 5. Onglet TOTAUX ----------
    ws = wb.create_sheet(title="TOTAUX")
    ws.cell(row=1, column=5, value=f"SOMME PERIODE").font = bold_font

    # Légende couleurs (demande user msg 181)
    ws.cell(row=1, column=13, value="LÉGENDE :").font = bold_font
    c = ws.cell(row=1, column=14, value="  Auto-rempli  ")
    c.fill = auto_fill; c.border = border; c.alignment = center
    c = ws.cell(row=1, column=15, value="  À remplir  ")
    c.fill = manual_fill; c.border = border; c.alignment = center
    c = ws.cell(row=1, column=16, value="  Règlement  ")
    c.fill = gray_fill; c.border = border; c.alignment = center
    c = ws.cell(row=1, column=17, value="  Facturation ext.  ")
    c.fill = beige_fill; c.border = border; c.alignment = center
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 14
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 20

    tot_headers = ["MOIS", "ESPECES", "CHEQUES", "CB", "PNF", "TOTAL CA",
                   "ENCAISSEM", "DEPENSES", "SOLDE CAISSE", "CHQ+REM", "CB+REM"]
    for col_i, h in enumerate(tot_headers, start=1):
        c = ws.cell(row=3, column=col_i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    for i, (month_key, sname, total_row) in enumerate(month_totals_by_sheet):
        r = 5 + i
        y, m = map(int, month_key.split("-"))
        ws.cell(row=r, column=1, value=f"{MONTH_ABBR[m]} {y}")
        for col_i, col_letter in enumerate(["B", "C", "D", "E", "F", "G", "H", "I", "L", "M"], start=2):
            ws.cell(row=r, column=col_i, value=f"='{sname}'!{col_letter}{total_row}")
        # Bordures + fond vert pâle (toutes les cases du tableau récap sont auto-calculées)
        for col_i in range(1, 12):
            cell = ws.cell(row=r, column=col_i)
            cell.border = border
            cell.fill = auto_fill
            if col_i > 1:
                cell.number_format = "#,##0.00"

    # Ligne TOTAUX globale
    if month_totals_by_sheet:
        last = 5 + len(month_totals_by_sheet) - 1
        tot_row = last + 2
        ws.cell(row=tot_row, column=1, value="TOTAUX :").font = bold_font
        for col_i in range(2, 12):
            letter = openpyxl.utils.get_column_letter(col_i)
            cell = ws.cell(row=tot_row, column=col_i, value=f"=SUM({letter}5:{letter}{last})")
            cell.font = bold_font
            cell.number_format = "#,##0.00"
        # Bordures + fond gris clair sur la ligne totaux globale
        for col_i in range(1, 12):
            cell = ws.cell(row=tot_row, column=col_i)
            cell.border = border
            cell.fill = total_fill

        # TOTAL CA global
        ws.cell(row=tot_row + 2, column=5, value="TOTAL CA :").font = bold_font
        c = ws.cell(row=tot_row + 2, column=6, value=f"=F{tot_row}+I{tot_row}+J{tot_row}+L{tot_row}")
        c.font = bold_font
        c.number_format = "#,##0.00"
        for col_i in (5, 6):
            cell = ws.cell(row=tot_row + 2, column=col_i)
            cell.border = border
            cell.fill = total_fill

    for col_letter, w in {"A": 12, "B": 11, "C": 11, "D": 11, "E": 10, "F": 12,
                          "G": 12, "H": 12, "I": 14, "J": 12, "K": 12}.items():
        ws.column_dimensions[col_letter].width = w

    # Pas de données -> on crée quand même un onglet courant
    if not ordered_month_keys:
        now = datetime.now()
        sname = f"{MONTH_ABBR[now.month]} {str(now.year)[-2:]}"
        wb.remove(ws)
        ws2 = wb.create_sheet(title=sname)
        for col_i, h in enumerate(HEADERS_ROW1, start=1):
            c = ws2.cell(row=1, column=col_i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
        ws2.merge_cells("J1:K1")
        ws2.merge_cells("L1:M1")
        for col_letter, w in COL_WIDTHS.items():
            ws2.column_dimensions[col_letter].width = w
        wb.create_sheet(title="TOTAUX")

    # ---------- 6. Sortie ----------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"journal_caisse_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def seed_admin_user():
    """Idempotent : crée le compte admin s'il n'existe pas encore."""
    admin_email = (os.environ.get("ADMIN_EMAIL") or "").strip().lower()
    admin_password = os.environ.get("ADMIN_PASSWORD") or ""
    if not admin_email or not admin_password:
        logger.warning("ADMIN_EMAIL / ADMIN_PASSWORD non configurés — pas de seed")
        return
    existing = await db.users.find_one({"email": admin_email})
    if existing:
        logger.info(f"Admin already seeded: {admin_email}")
        return
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": admin_email,
        "name": "Administrateur DCLIC",
        "password_hash": _hash_password(admin_password),
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    logger.info(f"Admin seeded: {admin_email}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
